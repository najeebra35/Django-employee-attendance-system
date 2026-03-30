"""
AttendPro AI Chat — ai_views.py
Handles the chat page and all AI agent logic using Google Gemini (free).
"""

import json
import datetime
import io
import os
import urllib.request
import urllib.error
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.conf import settings
from django.db.models import Q, Count, Sum
from attendance_app.models import (
    Employee, Attendance, Holiday, LeaveRequest,
    EmployeeDocument, CompanySettings
)


# ── Gemini API call ────────────────────────────────────────────────────────────

# ── Simple in-memory cache ─────────────────────────────────────────────────────
_response_cache = {}
_cache_timestamps = {}
CACHE_TTL_SECONDS = 120


def call_ai(prompt):
    """Call Groq API (free, works in UAE). Falls back to Gemini if no Groq key."""
    import hashlib, time

    cache_key = hashlib.md5(prompt.encode()).hexdigest()
    now = time.time()
    if cache_key in _response_cache:
        if now - _cache_timestamps.get(cache_key, 0) < CACHE_TTL_SECONDS:
            return _response_cache[cache_key]

    if len(prompt) > 5000:
        prompt = prompt[:5000] + "\n[truncated]"

    # Check keys in priority order: OpenRouter → Groq (legacy) → Gemini
    openrouter_key = getattr(settings, 'OPENROUTER_API_KEY', '')
    groq_key       = getattr(settings, 'GROQ_API_KEY', '')
    gemini_key     = getattr(settings, 'GEMINI_API_KEY', '')

    active_key = None
    if openrouter_key and openrouter_key not in ('YOUR_OPENROUTER_API_KEY_HERE', ''):
        active_key = openrouter_key
    elif groq_key and groq_key not in ('YOUR_GROQ_API_KEY_HERE', ''):
        active_key = groq_key  # works as OpenRouter key if user updates

    if active_key:
        result = _call_groq(prompt, active_key)  # _call_groq now calls OpenRouter
    elif gemini_key and gemini_key not in ('YOUR_GEMINI_API_KEY_HERE', ''):
        result = _call_gemini(prompt, gemini_key)
    else:
        return (
            "❌ No API key configured.\n\n"
            "Please set OPENROUTER_API_KEY in settings.py\n"
            "Get free key at: https://openrouter.ai/keys"
        )

    if result and not result.startswith("❌") and not result.startswith("⏳"):
        _response_cache[cache_key] = result
        _cache_timestamps[cache_key] = time.time()
    return result


def _call_groq(prompt, api_key):
    """OpenRouter API — free tier, works worldwide including UAE."""
    url = "https://openrouter.ai/api/v1/chat/completions"
    # Free models on OpenRouter — tries each in order if previous not available
    free_models = [
        "meta-llama/llama-4-scout:free",
        "meta-llama/llama-3.1-8b-instruct:free",
        "meta-llama/llama-3-8b-instruct:free",
        "mistralai/mistral-7b-instruct:free",
        "google/gemma-3-4b-it:free",
        "microsoft/phi-3-mini-128k-instruct:free",
    ]

    last_error = "❌ No free models available on OpenRouter. Please try again later."

    for model in free_models:
        payload = json.dumps({
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.2,
            "max_tokens": 1024,
        }).encode("utf-8")
        req = urllib.request.Request(
            url, data=payload,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
                "HTTP-Referer": "https://attendpro.app",
                "X-Title": "AttendPro AI",
            },
            method="POST"
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                choices = data.get("choices", [])
                if choices and choices[0]["message"]["content"].strip():
                    return choices[0]["message"]["content"]
                if "error" in data:
                    err = data["error"]
                    code = err.get("code", 0)
                    if code == 404:
                        last_error = f"model {model} not found"
                        continue
                    return f"❌ API error: {err.get('message', str(err))}"
                last_error = "❌ Empty response. Trying next model..."
                continue
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8")
            try:
                err_data = json.loads(body)
                err_msg = err_data.get("error", {}).get("message", body[:200])
            except Exception:
                err_msg = body[:200]
            if e.code == 401:
                return "❌ OpenRouter key invalid. Please check: https://openrouter.ai/keys"
            if e.code == 402:
                return "❌ OpenRouter requires credits for this model. Try a different free model."
            if e.code == 404:
                last_error = f"model {model} not found (404)"
                continue  # try next model
            if e.code == 429:
                return "⏳ Too many requests. Please wait 30 seconds and try again."
            last_error = f"❌ API error {e.code}: {err_msg}"
            continue
        except Exception as ex:
            last_error = f"❌ Connection error: {str(ex)}"
            continue

    return last_error


def _call_gemini(prompt, api_key):
    """Gemini fallback."""
    models = ["gemini-2.0-flash", "gemini-1.5-flash-latest", "gemini-1.5-flash"]
    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.2, "maxOutputTokens": 1024},
    }).encode("utf-8")
    for model in models:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
        req = urllib.request.Request(url, data=payload,
            headers={"Content-Type": "application/json"}, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                return data["candidates"][0]["content"]["parts"][0]["text"]
        except urllib.error.HTTPError as e:
            if e.code == 404:
                continue
            body = e.read().decode("utf-8")
            if e.code == 429:
                return "❌ Gemini quota exceeded. Gemini free tier is blocked in UAE. Use Groq instead."
            return f"❌ Gemini error {e.code}: {body[:200]}"
        except Exception as ex:
            return f"❌ Error: {str(ex)}"
    return "❌ All Gemini models unavailable."


# keep old name as alias so nothing breaks
def call_gemini(prompt, system_prompt=""):
    full = f"{system_prompt}\n\n{prompt}" if system_prompt else prompt
    return call_ai(full)


# ── Data fetcher functions ─────────────────────────────────────────────────────

def get_today_attendance():
    today = timezone.localdate()
    records = Attendance.objects.filter(date=today).select_related('employee')
    result = {
        "date": str(today),
        "present": [],
        "absent": [],
        "half_day": [],
        "on_leave": [],
        "holiday": [],
        "not_marked": [],
    }

    marked_ids = set()
    for r in records:
        marked_ids.add(r.employee_id)
        entry = {
            "name": r.employee.name,
            "job_title": r.employee.job_title or "",
            "in_time": str(r.in_time) if r.in_time else "",
            "out_time": str(r.out_time) if r.out_time else "",
            "ot_hours": float(r.ot_hours),
            "notes": r.notes or "",
        }
        if r.status == "present":
            result["present"].append(entry)
        elif r.status == "absent":
            result["absent"].append(entry)
        elif r.status == "half_day":
            result["half_day"].append(entry)
        elif r.status == "leave":
            result["on_leave"].append(entry)
        elif r.status == "holiday":
            result["holiday"].append(entry)

    # Employees with no attendance marked today
    all_active = Employee.objects.filter(status='active')
    for emp in all_active:
        if emp.id not in marked_ids:
            result["not_marked"].append({"name": emp.name, "job_title": emp.job_title or ""})

    result["summary"] = {
        "present_count": len(result["present"]),
        "absent_count": len(result["absent"]),
        "half_day_count": len(result["half_day"]),
        "on_leave_count": len(result["on_leave"]),
        "not_marked_count": len(result["not_marked"]),
        "total_employees": all_active.count(),
    }
    return result


def get_monthly_attendance(year=None, month=None):
    today = timezone.localdate()
    year  = year  or today.year
    month = month or today.month
    month_name = datetime.date(year, month, 1).strftime("%B %Y")

    employees = Employee.objects.filter(status='active').order_by('name')
    result = {"period": month_name, "year": year, "month": month, "employees": []}

    for emp in employees:
        records = Attendance.objects.filter(
            employee=emp, date__year=year, date__month=month
        )
        counts = {"present": 0, "absent": 0, "half_day": 0, "leave": 0, "holiday": 0}
        total_ot = 0
        for r in records:
            counts[r.status] = counts.get(r.status, 0) + 1
            total_ot += float(r.ot_hours)

        total_days = counts["present"] + counts["absent"] + counts["half_day"] + counts["leave"]
        att_rate = round((counts["present"] + counts["half_day"] * 0.5) / total_days * 100, 1) if total_days > 0 else 0

        result["employees"].append({
            "name": emp.name,
            "job_title": emp.job_title or "",
            "present": counts["present"],
            "absent": counts["absent"],
            "half_day": counts["half_day"],
            "on_leave": counts["leave"],
            "holiday": counts["holiday"],
            "total_ot_hours": round(total_ot, 2),
            "attendance_rate": att_rate,
        })
    return result


def get_employee_attendance(name_query, year=None, month=None):
    today = timezone.localdate()
    year  = year  or today.year
    month = month or today.month

    emp = Employee.objects.filter(
        name__icontains=name_query, status='active'
    ).first()
    if not emp:
        return {"error": f"No active employee found matching '{name_query}'"}

    month_name = datetime.date(year, month, 1).strftime("%B %Y")
    records = Attendance.objects.filter(
        employee=emp, date__year=year, date__month=month
    ).order_by('date')

    days = []
    total_ot = 0
    counts = {"present": 0, "absent": 0, "half_day": 0, "leave": 0, "holiday": 0}
    for r in records:
        total_ot += float(r.ot_hours)
        counts[r.status] = counts.get(r.status, 0) + 1
        days.append({
            "date": str(r.date),
            "day": r.date.strftime("%A"),
            "status": r.status,
            "in_time": str(r.in_time) if r.in_time else "",
            "out_time": str(r.out_time) if r.out_time else "",
            "ot_hours": float(r.ot_hours),
            "notes": r.notes or "",
        })

    total_days = counts["present"] + counts["absent"] + counts["half_day"] + counts["leave"]
    att_rate = round((counts["present"] + counts["half_day"] * 0.5) / total_days * 100, 1) if total_days > 0 else 0

    return {
        "employee": emp.name,
        "job_title": emp.job_title or "",
        "period": month_name,
        "summary": {
            "present": counts["present"],
            "absent": counts["absent"],
            "half_day": counts["half_day"],
            "on_leave": counts["leave"],
            "total_ot_hours": round(total_ot, 2),
            "attendance_rate": att_rate,
        },
        "daily_records": days,
    }


def get_expiring_documents(days=60):
    today = timezone.localdate()
    docs = EmployeeDocument.objects.select_related('employee', 'doc_type').filter(
        employee__status='active', expiry_date__isnull=False
    ).order_by('expiry_date')

    result = {"expired": [], "expiring_soon": []}
    for d in docs:
        if d.expiry_date < today:
            result["expired"].append({
                "employee": d.employee.name,
                "document": d.doc_type.name,
                "doc_number": d.doc_number or "",
                "expiry_date": str(d.expiry_date),
                "days_overdue": (today - d.expiry_date).days,
            })
        elif (d.expiry_date - today).days <= days:
            result["expiring_soon"].append({
                "employee": d.employee.name,
                "document": d.doc_type.name,
                "doc_number": d.doc_number or "",
                "expiry_date": str(d.expiry_date),
                "days_left": (d.expiry_date - today).days,
            })
    result["expired_count"]  = len(result["expired"])
    result["expiring_count"] = len(result["expiring_soon"])
    return result


def get_late_arrivals(year=None, month=None):
    today = timezone.localdate()
    year  = year  or today.year
    month = month or today.month
    cs    = CompanySettings.get_settings()
    std   = cs.default_in_time

    records = Attendance.objects.filter(
        date__year=year, date__month=month,
        status='present', in_time__isnull=False
    ).select_related('employee').order_by('employee__name')

    result = []
    for r in records:
        if r.in_time > std:
            diff_mins = (
                datetime.datetime.combine(r.date, r.in_time) -
                datetime.datetime.combine(r.date, std)
            ).seconds // 60
            result.append({
                "employee": r.employee.name,
                "date": str(r.date),
                "in_time": str(r.in_time),
                "standard_time": str(std),
                "late_by_minutes": diff_mins,
            })

    return {
        "period": datetime.date(year, month, 1).strftime("%B %Y"),
        "standard_in_time": str(std),
        "late_records": result,
        "total_late_instances": len(result),
    }


# ── Excel export helper ────────────────────────────────────────────────────────

def generate_excel_from_data(data, title="Report"):
    """Generate Excel from structured data dict, return bytes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="E8650A")
    center = Alignment(horizontal="center")

    row = 1
    ws.cell(row, 1, title).font = Font(bold=True, size=13)
    ws.cell(row, 1).alignment = center
    row += 1

    if "employees" in data:
        # Monthly summary format
        headers = ["Name", "Job Title", "Present", "Absent", "Half Day", "On Leave", "OT Hours", "Att. Rate %"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.font = hfont; c.fill = hfill; c.alignment = center
        row += 1
        for emp in data["employees"]:
            ws.cell(row, 1, emp["name"])
            ws.cell(row, 2, emp.get("job_title", ""))
            ws.cell(row, 3, emp.get("present", 0))
            ws.cell(row, 4, emp.get("absent", 0))
            ws.cell(row, 5, emp.get("half_day", 0))
            ws.cell(row, 6, emp.get("on_leave", 0))
            ws.cell(row, 7, emp.get("total_ot_hours", 0))
            ws.cell(row, 8, emp.get("attendance_rate", 0))
            row += 1

    elif "daily_records" in data:
        # Employee detail format
        ws.cell(row, 1, f"Employee: {data['employee']}").font = Font(bold=True)
        ws.cell(row, 3, f"Period: {data['period']}").font = Font(bold=True)
        row += 1
        s = data.get("summary", {})
        ws.cell(row, 1, f"Present: {s.get('present',0)}  Absent: {s.get('absent',0)}  Half Day: {s.get('half_day',0)}  OT: {s.get('total_ot_hours',0)}h")
        row += 1

        headers = ["Date", "Day", "Status", "In Time", "Out Time", "OT Hours", "Notes"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.font = hfont; c.fill = hfill; c.alignment = center
        row += 1
        for d in data["daily_records"]:
            ws.cell(row, 1, d["date"])
            ws.cell(row, 2, d["day"])
            ws.cell(row, 3, d["status"].replace("_", " ").title())
            ws.cell(row, 4, d["in_time"])
            ws.cell(row, 5, d["out_time"])
            ws.cell(row, 6, d["ot_hours"])
            ws.cell(row, 7, d["notes"])
            row += 1

    elif "present" in data and "absent" in data and "date" in data:
        # Today's attendance
        ws.cell(row, 1, f"Date: {data['date']}  Present: {data['summary']['present_count']}  Absent: {data['summary']['absent_count']}  Not Marked: {data['summary']['not_marked_count']}").font = Font(bold=True)
        row += 1
        headers = ["Name", "Job Title", "Status", "In Time", "Out Time", "OT Hours"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.font = hfont; c.fill = hfill; c.alignment = center
        row += 1
        for status_key, label in [("present","Present"),("absent","Absent"),("half_day","Half Day"),("on_leave","On Leave"),("not_marked","Not Marked")]:
            for emp in data.get(status_key, []):
                ws.cell(row, 1, emp["name"])
                ws.cell(row, 2, emp.get("job_title",""))
                ws.cell(row, 3, label)
                ws.cell(row, 4, emp.get("in_time",""))
                ws.cell(row, 5, emp.get("out_time",""))
                ws.cell(row, 6, emp.get("ot_hours",""))
                row += 1

    elif "expired" in data or "expiring_soon" in data:
        # Documents
        headers = ["Employee", "Document", "Doc Number", "Expiry Date", "Status", "Days"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.font = hfont; c.fill = hfill; c.alignment = center
        row += 1
        for d in data.get("expired", []):
            ws.append([d["employee"], d["document"], d["doc_number"], d["expiry_date"], "EXPIRED", f"-{d['days_overdue']} days"])
        for d in data.get("expiring_soon", []):
            ws.append([d["employee"], d["document"], d["doc_number"], d["expiry_date"], "EXPIRING", f"{d['days_left']} days left"])

    elif "late_records" in data:
        headers = ["Employee", "Date", "In Time", "Standard Time", "Late By (mins)"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.font = hfont; c.fill = hfill; c.alignment = center
        row += 1
        for d in data["late_records"]:
            ws.append([d["employee"], d["date"], d["in_time"], d["standard_time"], d["late_by_minutes"]])

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── System prompt for Gemini ───────────────────────────────────────────────────

SYSTEM_PROMPT = """AttendPro AI. Date: {today}. Answer attendance questions using the data below. Be friendly and concise. Show totals first, then details. Use simple text lists/tables.

DATA:
{data}

Q: {question}
A:"""


# ── Intent detection ───────────────────────────────────────────────────────────

def detect_intent_and_fetch(question):
    """
    Detect what the user is asking and fetch relevant data from DB.
    Returns (data_dict, data_label, export_filename)
    """
    q = question.lower()
    today = timezone.localdate()

    # Parse month/year from question if mentioned
    target_month = today.month
    target_year  = today.year

    month_names = {
        "january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
        "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
        "jan":1,"feb":2,"mar":3,"apr":4,"jun":6,"jul":7,"aug":8,
        "sep":9,"oct":10,"nov":11,"dec":12
    }
    for mname, mnum in month_names.items():
        if mname in q:
            target_month = mnum
            break

    # Parse year
    import re
    year_match = re.search(r'20\d\d', q)
    if year_match:
        target_year = int(year_match.group())

    # Detect "last month"
    if "last month" in q or "previous month" in q:
        d = today.replace(day=1) - datetime.timedelta(days=1)
        target_month = d.month
        target_year  = d.year

    # ── Route to correct data ──────────────────────────────────────────────────

    # Document queries
    if any(w in q for w in ["document", "visa", "passport", "emirates id", "expir", "id card", "labour card"]):
        data = get_expiring_documents()
        return data, "expiring_documents", "document_expiry_report.xlsx"

    # Late arrival queries
    if any(w in q for w in ["late", "delay", "early"]):
        data = get_late_arrivals(target_year, target_month)
        return data, "late_arrivals", f"late_arrivals_{target_year}_{target_month:02d}.xlsx"

    # Today's attendance
    if any(w in q for w in ["today", "now", "current", "right now", "this day"]):
        data = get_today_attendance()
        return data, "today_attendance", f"attendance_today_{today}.xlsx"

    # Specific employee query — check if any employee name is mentioned
    active_employees = list(Employee.objects.filter(status='active').values_list('name', flat=True))
    matched_emp = None
    for emp_name in active_employees:
        # Check if any word in employee name appears in question
        for part in emp_name.lower().split():
            if len(part) > 2 and part in q:
                matched_emp = emp_name
                break
        if matched_emp:
            break

    if matched_emp:
        data = get_employee_attendance(matched_emp, target_year, target_month)
        safe_name = matched_emp.replace(" ", "_").lower()
        return data, "employee_attendance", f"{safe_name}_{target_year}_{target_month:02d}.xlsx"

    # Monthly attendance (default)
    data = get_monthly_attendance(target_year, target_month)
    return data, "monthly_attendance", f"monthly_attendance_{target_year}_{target_month:02d}.xlsx"


# ── Main chat view ─────────────────────────────────────────────────────────────

@login_required
def ai_chat_page(request):
    """Render the chat page."""
    return render(request, 'attendance_app/ai_chat.html')


@login_required
def ai_chat_message(request):
    """Handle incoming chat message, return AI response."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        body    = json.loads(request.body)
        question = body.get('message', '').strip()
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    if not question:
        return JsonResponse({'error': 'Empty message'}, status=400)

    today  = timezone.localdate()
    export = body.get('export', False)

    # Fetch relevant data
    data, data_label, export_filename = detect_intent_and_fetch(question)

    # If export requested — generate Excel and return download token
    if export or any(w in question.lower() for w in ["export", "download", "excel", "file", "send"]):
        excel_bytes = generate_excel_from_data(data, data_label.replace("_", " ").title())
        if excel_bytes:
            # Store in session for download
            import base64
            request.session['ai_export_data']     = base64.b64encode(excel_bytes).decode()
            request.session['ai_export_filename']  = export_filename
            request.session.modified = True

    # Build compact data summary (much smaller than full JSON — saves tokens & avoids rate limit)
    month_name = datetime.date(today.year, today.month, 1).strftime("%B %Y")

    def compact_data(d, label):
        """Convert data dict to short readable text instead of full JSON."""
        lines = []
        if label == "today_attendance":
            s = d.get("summary", {})
            lines.append(f"Date: {d.get('date')} | Present:{s.get('present_count',0)} Absent:{s.get('absent_count',0)} HalfDay:{s.get('not_marked_count',0)} NotMarked:{s.get('not_marked_count',0)} Total:{s.get('total_employees',0)}")
            if d.get("present"):
                lines.append("PRESENT: " + ", ".join(e["name"] + (f"(in:{e['in_time']})" if e.get("in_time") else "") for e in d["present"]))
            if d.get("absent"):
                lines.append("ABSENT: " + ", ".join(e["name"] for e in d["absent"]))
            if d.get("half_day"):
                lines.append("HALF DAY: " + ", ".join(e["name"] for e in d["half_day"]))
            if d.get("on_leave"):
                lines.append("ON LEAVE: " + ", ".join(e["name"] for e in d["on_leave"]))
            if d.get("not_marked"):
                lines.append("NOT MARKED: " + ", ".join(e["name"] for e in d["not_marked"]))

        elif label == "monthly_attendance":
            lines.append(f"Period: {d.get('period')} | {len(d.get('employees',[]))} employees")
            for e in d.get("employees", []):
                lines.append(f"  {e['name']}: Present={e['present']} Absent={e['absent']} HalfDay={e['half_day']} Leave={e['on_leave']} OT={e['total_ot_hours']}h Rate={e['attendance_rate']}%")

        elif label == "employee_attendance":
            if "error" in d:
                return d["error"]
            s = d.get("summary", {})
            lines.append(f"Employee: {d.get('employee')} | Period: {d.get('period')}")
            lines.append(f"Present={s.get('present',0)} Absent={s.get('absent',0)} HalfDay={s.get('half_day',0)} Leave={s.get('on_leave',0)} OT={s.get('total_ot_hours',0)}h Rate={s.get('attendance_rate',0)}%")
            for r in d.get("daily_records", []):
                st = r["status"].replace("_"," ").upper()
                t = f" {r['in_time']}-{r['out_time']}" if r.get("in_time") else ""
                ot = f" OT:{r['ot_hours']}h" if r.get("ot_hours") else ""
                lines.append(f"  {r['date']} {r['day'][:3]}: {st}{t}{ot}")

        elif label == "expiring_documents":
            lines.append(f"Expired: {d.get('expired_count',0)} | Expiring soon: {d.get('expiring_count',0)}")
            for x in d.get("expired", []):
                lines.append(f"  EXPIRED: {x['employee']} - {x['document']} (expired {x['expiry_date']}, {x['days_overdue']}d ago)")
            for x in d.get("expiring_soon", []):
                lines.append(f"  EXPIRING: {x['employee']} - {x['document']} (expires {x['expiry_date']}, {x['days_left']}d left)")

        elif label == "late_arrivals":
            lines.append(f"Period: {d.get('period')} | Standard time: {d.get('standard_in_time')} | Total late instances: {d.get('total_late_instances',0)}")
            for r in d.get("late_records", []):
                lines.append(f"  {r['employee']} on {r['date']}: arrived {r['in_time']} (late by {r['late_by_minutes']} min)")

        return "\n".join(lines)

    compact = compact_data(data, data_label)
    prompt_filled = SYSTEM_PROMPT.format(
        today=str(today),
        data=compact,
        question=question,
    )

    # Call Gemini
    ai_response = call_ai(prompt_filled)

    has_export = 'ai_export_data' in request.session

    return JsonResponse({
        'reply':       ai_response,
        'data_type':   data_label,
        'has_export':  has_export,
        'export_name': export_filename if has_export else '',
    })


@login_required
def ai_chat_export(request):
    """Download the last exported Excel file."""
    import base64
    excel_b64 = request.session.pop('ai_export_data', None)
    filename  = request.session.pop('ai_export_filename', 'export.xlsx')
    request.session.modified = True

    if not excel_b64:
        from django.shortcuts import redirect
        return redirect('ai_chat')

    excel_bytes = base64.b64decode(excel_b64)
    resp = HttpResponse(
        excel_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
