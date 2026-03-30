import os

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Q, Count
from django.conf import settings
from .models import (Employee, Attendance, Holiday, LeaveType, LeaveRequest,
                     UserPermission, ActivityLog, CompanySettings,
                     DocumentType, EmployeeDocument)
from .middleware import log_activity
from .decorators import permission_required
import datetime
import json
import calendar
import io

# ─── AUTH ────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            log_activity(user, 'login', 'Auth', description=f'{user.username} logged in', request=request)
            # If user is linked to an employee (portal user) and not admin → go to portal
            try:
                if not user.is_superuser and user.employee_profile:
                    return redirect('portal_dashboard')
            except Exception:
                pass
            return redirect('dashboard')
        messages.error(request, 'Invalid username or password.')
    return render(request, 'attendance_app/login.html')


def logout_view(request):
    if request.user.is_authenticated:
        log_activity(request.user, 'logout', 'Auth', description=f'{request.user.username} logged out', request=request)
    logout(request)
    return redirect('login')


# ─── PERMISSION HELPER ────────────────────────────────────────────────────────

def has_perm(user, perm):
    if user.is_superuser:
        return True
    try:
        up = UserPermission.objects.get(user=user)
        return perm in up.permissions
    except UserPermission.DoesNotExist:
        return False


# ─── DASHBOARD ───────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    if not has_perm(request.user, 'dashboard_view') and not request.user.is_superuser:
        return redirect_to_first_permitted(request)

    today = timezone.localdate()
    total_employees = Employee.objects.filter(status="active").count()
    today_atts = list(Attendance.objects.filter(date=today).select_related("employee"))
    today_attendance = sum(1 for a in today_atts if a.status == "present")
    today_absent     = sum(1 for a in today_atts if a.status == "absent")
    today_half       = sum(1 for a in today_atts if a.status == "half_day")
    today_leave      = sum(1 for a in today_atts if a.status == "leave")
    today_total_marked = len(today_atts)
    today_not_marked = max(0, total_employees - today_total_marked)

    upcoming_holidays = Holiday.objects.filter(date__gte=today).order_by("date")[:5]
    recent_activity   = ActivityLog.objects.select_related("user").all()[:8]
    pending_leaves    = LeaveRequest.objects.filter(status="pending").count()
    approved_leaves   = LeaveRequest.objects.filter(
        status="approved", start_date__lte=today, end_date__gte=today
    ).count()

    # ── Month summary ──────────────────────────────────────────────────────
    month_start = today.replace(day=1)
    month_atts = Attendance.objects.filter(date__gte=month_start, date__lte=today)
    month_present  = month_atts.filter(status="present").count()
    month_absent   = month_atts.filter(status="absent").count()
    month_leave    = month_atts.filter(status="leave").count()
    month_half     = month_atts.filter(status="half_day").count()

    # ── 30-day trend data for chart ────────────────────────────────────────
    trend_labels, trend_present, trend_absent = [], [], []
    for i in range(29, -1, -1):
        d = today - datetime.timedelta(days=i)
        day_atts = Attendance.objects.filter(date=d)
        trend_labels.append(d.strftime("%d %b"))
        trend_present.append(day_atts.filter(status="present").count())
        trend_absent.append(day_atts.filter(status="absent").count())

    # ── Top employees this month (most present days) ───────────────────────
    from django.db.models import Count as DCount
    top_employees = (
        Attendance.objects.filter(date__gte=month_start, date__lte=today, status="present")
        .values("employee__name")
        .annotate(present_days=DCount("id"))
        .order_by("-present_days")[:5]
    )

    # ── Today absent employees ─────────────────────────────────────────────
    today_absent_emps = [a.employee for a in today_atts if a.status == "absent"][:6]

    # ── Next 7 days events ─────────────────────────────────────────────────
    week_end = today + datetime.timedelta(days=7)
    week_holidays = Holiday.objects.filter(date__gte=today, date__lte=week_end).order_by("date")

    context = {
        "total_employees": total_employees,
        "today_attendance": today_attendance,
        "today_absent": today_absent,
        "today_half": today_half,
        "today_leave": today_leave,
        "today_total_marked": today_total_marked,
        "today_not_marked": today_not_marked,
        "upcoming_holidays": upcoming_holidays,
        "week_holidays": week_holidays,
        "recent_activity": recent_activity,
        "pending_leaves": pending_leaves,
        "approved_leaves": approved_leaves,
        "today": today,
        "month_start": month_start,
        "month_present": month_present,
        "month_absent": month_absent,
        "month_leave": month_leave,
        "month_half": month_half,
        "trend_labels": trend_labels,
        "trend_present": trend_present,
        "trend_absent": trend_absent,
        "top_employees": list(top_employees),
        "today_absent_emps": today_absent_emps,
    }
    return render(request, "attendance_app/dashboard.html", context)


def redirect_to_first_permitted(request):
    perms_pages = [
        ('employee_view', 'employee_list'),
        ('attendance_view', 'attendance_list'),
        ('holiday_view', 'holiday_list'),
        ('leave_view', 'leave_list'),
        ('export_view', 'export_attendance'),
        ('activity_view', 'activity_log'),
    ]
    for perm, url_name in perms_pages:
        if has_perm(request.user, perm):
            return redirect(url_name)
    return render(request, 'attendance_app/no_access.html')


# ─── EMPLOYEES ───────────────────────────────────────────────────────────────

@login_required
def employee_list(request):
    if not has_perm(request.user, 'employee_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    employees = Employee.objects.all()
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    type_filter = request.GET.get('type', '')

    if search:
        employees = employees.filter(
            Q(name__icontains=search) | Q(job_title__icontains=search) |
            Q(emirates_id__icontains=search) | Q(mobile__icontains=search)
        )
    if status_filter:
        employees = employees.filter(status=status_filter)
    if type_filter:
        employees = employees.filter(emp_type=type_filter)

    return render(request, 'attendance_app/employee_list.html', {
        'employees': employees,
        'search': search,
        'status_filter': status_filter,
        'type_filter': type_filter,
    })


@login_required
def employee_add(request):
    if not has_perm(request.user, 'employee_add'):
        messages.error(request, 'Access denied.')
        return redirect('employee_list')

    if request.method == 'POST':
        emp = Employee()
        emp.name = request.POST.get('name', '').strip()
        if not emp.name:
            messages.error(request, 'Employee name is required.')
            return render(request, 'attendance_app/employee_form.html', {'form_data': request.POST})

        emp.emirates_id = request.POST.get('emirates_id') or None
        emp.mobile = request.POST.get('mobile') or None
        emp.dob = request.POST.get('dob') or None
        emp.joining_date = request.POST.get('joining_date') or None
        emp.job_title = request.POST.get('job_title') or None
        emp.country = request.POST.get('country') or None
        emp.address = request.POST.get('address') or None
        emp.description = request.POST.get('description') or None
        emp.emp_type = request.POST.get('emp_type', 'permanent')
        emp.status = request.POST.get('status', 'active')
        if request.FILES.get('photo'):
            emp.photo = request.FILES['photo']
        emp.save()

        log_activity(request.user, 'create', 'Employee', emp, f'Added employee {emp.name}', request)
        messages.success(request, f'Employee {emp.name} added successfully.')
        return redirect('employee_list')

    return render(request, 'attendance_app/employee_form.html', {'action': 'Add'})


@login_required
def employee_detail(request, pk):
    if not has_perm(request.user, 'employee_detail'):
        messages.error(request, 'Access denied.')
        return redirect('employee_list')

    employee = get_object_or_404(Employee, pk=pk)
    recent_attendance = Attendance.objects.filter(employee=employee).order_by('-date')[:30]
    leave_requests = LeaveRequest.objects.filter(employee=employee).order_by('-applied_on')[:10]
    activities = ActivityLog.objects.filter(
        description__icontains=employee.name
    ).order_by('-timestamp')[:30]

    # Stats
    today = timezone.localdate()
    month_start = today.replace(day=1)
    month_present = Attendance.objects.filter(employee=employee, date__gte=month_start, status='present').count()
    month_absent = Attendance.objects.filter(employee=employee, date__gte=month_start, status='absent').count()
    month_leave = Attendance.objects.filter(employee=employee, date__gte=month_start, status='leave').count()

    return render(request, 'attendance_app/employee_detail.html', {
        'employee': employee,
        'recent_attendance': recent_attendance,
        'leave_requests': leave_requests,
        'activities': activities,
        'month_present': month_present,
        'month_absent': month_absent,
        'month_leave': month_leave,
    })


@login_required
def employee_edit(request, pk):
    if not has_perm(request.user, 'employee_edit'):
        messages.error(request, 'Access denied.')
        return redirect('employee_list')

    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.name = request.POST.get('name', '').strip()
        if not employee.name:
            messages.error(request, 'Employee name is required.')
            return render(request, 'attendance_app/employee_form.html', {'employee': employee, 'action': 'Edit'})

        employee.emirates_id = request.POST.get('emirates_id') or None
        employee.mobile = request.POST.get('mobile') or None
        employee.dob = request.POST.get('dob') or None
        employee.joining_date = request.POST.get('joining_date') or None
        employee.job_title = request.POST.get('job_title') or None
        employee.country = request.POST.get('country') or None
        employee.address = request.POST.get('address') or None
        employee.description = request.POST.get('description') or None
        employee.emp_type = request.POST.get('emp_type', 'permanent')
        employee.status = request.POST.get('status', 'active')
        if request.FILES.get('photo'):
            employee.photo = request.FILES['photo']
        employee.save()

        log_activity(request.user, 'update', 'Employee', employee, f'Updated employee {employee.name}', request)
        messages.success(request, f'Employee {employee.name} updated.')
        return redirect('employee_detail', pk=pk)

    return render(request, 'attendance_app/employee_form.html', {'employee': employee, 'action': 'Edit'})


@login_required
def employee_delete(request, pk):
    if not has_perm(request.user, 'employee_delete'):
        messages.error(request, 'Access denied.')
        return redirect('employee_list')

    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        name = employee.name
        employee.delete()
        log_activity(request.user, 'delete', 'Employee', description=f'Deleted employee {name}', request=request)
        messages.success(request, f'Employee {name} deleted.')
        return redirect('employee_list')
    return render(request, 'attendance_app/employee_confirm_delete.html', {'employee': employee})


@login_required
def employee_calendar_data(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    year = int(request.GET.get('year', timezone.localdate().year))
    month = int(request.GET.get('month', timezone.localdate().month))

    attendances = Attendance.objects.filter(
        employee=employee,
        date__year=year,
        date__month=month
    )
    holidays = Holiday.objects.filter(date__year=year, date__month=month)

    events = []
    for att in attendances:
        color = {'present': '#22c55e', 'absent': '#ef4444', 'half_day': '#f59e0b',
                 'holiday': '#6366f1', 'leave': '#8b5cf6'}.get(att.status, '#6b7280')
        events.append({
            'date': att.date.isoformat(),
            'status': att.get_status_display(),
            'in_time': att.in_time.strftime('%H:%M') if att.in_time else '-',
            'out_time': att.out_time.strftime('%H:%M') if att.out_time else '-',
            'ot_hours': float(att.ot_hours),
            'total_hours': att.total_hours(),
            'color': color,
        })
    for hol in holidays:
        if not any(e['date'] == hol.date.isoformat() for e in events):
            events.append({
                'date': hol.date.isoformat(),
                'status': hol.name,
                'in_time': '-', 'out_time': '-', 'ot_hours': 0, 'total_hours': 0,
                'color': '#6366f1',
            })

    return JsonResponse({'events': events})


# ─── ATTENDANCE ───────────────────────────────────────────────────────────────

@login_required
def attendance_list(request):
    if not has_perm(request.user, 'attendance_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    today = timezone.localdate()
    date_str = request.GET.get('date', today.isoformat())
    try:
        selected_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        selected_date = today

    attendances = list(Attendance.objects.filter(date=selected_date).select_related('employee').order_by('employee__name'))
    is_holiday = Holiday.objects.filter(date=selected_date).first()

    stat_present = sum(1 for a in attendances if a.status == 'present')
    stat_absent  = sum(1 for a in attendances if a.status == 'absent')
    stat_half    = sum(1 for a in attendances if a.status == 'half_day')
    stat_leave   = sum(1 for a in attendances if a.status == 'leave')
    stat_total   = len(attendances)
    stat_total_hours = round(sum(float(a.total_hours()) for a in attendances), 1)
    stat_total_ot    = round(sum(float(a.ot_hours) for a in attendances), 1)
    attendance_rate  = round((stat_present / stat_total * 100) if stat_total else 0)

    month_qs = Attendance.objects.filter(
        date__year=selected_date.year,
        date__month=selected_date.month
    ).values("date").annotate(count=Count("id")).order_by("date")

    monthly_summary = [
        {"date": item["date"], "count": item["count"]}
        for item in month_qs
    ]

    return render(request, "attendance_app/attendance_list.html", {
        "attendances": attendances,
        "selected_date": selected_date,
        "is_holiday": is_holiday,
        "today": today,
        "yesterday": today - datetime.timedelta(days=1),
        "prev_date": selected_date - datetime.timedelta(days=1),
        "next_date": selected_date + datetime.timedelta(days=1),
        "stat_present": stat_present,
        "stat_absent": stat_absent,
        "stat_half": stat_half,
        "stat_leave": stat_leave,
        "stat_total": stat_total,
        "stat_total_hours": stat_total_hours,
        "stat_total_ot": stat_total_ot,
        "attendance_rate": attendance_rate,
        "monthly_summary": monthly_summary,
    })


@login_required
def attendance_mark(request):
    if not has_perm(request.user, 'attendance_add'):
        messages.error(request, 'Access denied.')
        return redirect('attendance_list')

    today = timezone.localdate()
    date_str = request.GET.get('date', today.isoformat())
    try:
        selected_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        selected_date = today

    active_employees = Employee.objects.filter(status='active').order_by('name')
    existing = {a.employee_id: a for a in Attendance.objects.filter(date=selected_date)}
    is_holiday = Holiday.objects.filter(date=selected_date).first()
    is_sunday = selected_date.weekday() == 6

    # Last working day data (for "copy previous day" feature)
    prev_date = selected_date - datetime.timedelta(days=1)
    prev_att = {a.employee_id: a for a in Attendance.objects.filter(date=prev_date)}

    employee_data = []
    for emp in active_employees:
        att = existing.get(emp.pk)
        prev = prev_att.get(emp.pk)
        employee_data.append({
            'employee': emp,
            'attendance': att,
            'prev_attendance': prev,
        })

    # Count already marked
    marked_count = len(existing)
    total_count = active_employees.count()

    return render(request, "attendance_app/attendance_mark.html", {
        "employee_data": employee_data,
        "selected_date": selected_date,
        "is_holiday": is_holiday,
        "is_sunday": is_sunday,
        "today": today,
        "prev_date": prev_date,
        "next_date": selected_date + datetime.timedelta(days=1),
        "marked_count": marked_count,
        "total_count": total_count,
    })


@login_required
@require_POST
def attendance_save(request):
    if not has_perm(request.user, 'attendance_add'):
        return JsonResponse({'error': 'Access denied'}, status=403)

    date_str = request.POST.get('date')
    selected_date = datetime.date.fromisoformat(date_str)
    active_employees = Employee.objects.filter(status='active')

    saved = 0
    for emp in active_employees:
        status = request.POST.get(f'status_{emp.pk}', 'absent')
        in_time_str = request.POST.get(f'in_time_{emp.pk}', '')
        out_time_str = request.POST.get(f'out_time_{emp.pk}', '')
        ot_hours = request.POST.get(f'ot_{emp.pk}', '0') or '0'

        att, created = Attendance.objects.get_or_create(
            employee=emp, date=selected_date,
            defaults={'created_by': request.user}
        )
        att.status = status
        att.in_time = datetime.time.fromisoformat(in_time_str) if in_time_str else None
        att.out_time = datetime.time.fromisoformat(out_time_str) if out_time_str else None
        try:
            att.ot_hours = float(ot_hours)
        except ValueError:
            att.ot_hours = 0
        att.save()
        saved += 1

    log_activity(request.user, 'create', 'Attendance', description=f'Marked attendance for {selected_date}', request=request)
    messages.success(request, f'Attendance saved for {selected_date} — {saved} employee(s) updated.')
    return redirect(f"/attendance/mark/?date={selected_date}")


@login_required
def attendance_edit(request, pk):
    if not has_perm(request.user, 'attendance_edit'):
        messages.error(request, 'Access denied.')
        return redirect('attendance_list')

    att = get_object_or_404(Attendance, pk=pk)
    if request.method == 'POST':
        att.status = request.POST.get('status', att.status)
        in_time_str = request.POST.get('in_time', '')
        out_time_str = request.POST.get('out_time', '')
        att.in_time = datetime.time.fromisoformat(in_time_str) if in_time_str else None
        att.out_time = datetime.time.fromisoformat(out_time_str) if out_time_str else None
        att.ot_hours = float(request.POST.get('ot_hours', 0) or 0)
        att.notes = request.POST.get('notes', '')
        att.save()
        log_activity(request.user, 'update', 'Attendance', att, f'Updated attendance for {att.employee.name} on {att.date}', request)
        messages.success(request, f'Attendance updated for {att.employee.name}.')
        return_url = request.POST.get('return_url', f'/attendance/?date={att.date}')
        return redirect(return_url)

    return render(request, 'attendance_app/attendance_edit.html', {
        'att': att,
        'return_url': request.GET.get('return_url', f'/attendance/?date={att.date}'),
    })


# ─── HOLIDAYS ─────────────────────────────────────────────────────────────────

@login_required
def holiday_list(request):
    if not has_perm(request.user, 'holiday_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    year = int(request.GET.get('year', timezone.localdate().year))
    holidays = Holiday.objects.filter(date__year=year).order_by('date')
    return render(request, 'attendance_app/holiday_list.html', {
        'holidays': holidays, 'year': year,
        'years': range(2020, timezone.localdate().year + 3),
    })


@login_required
def holiday_add(request):
    if not has_perm(request.user, 'holiday_add'):
        messages.error(request, 'Access denied.')
        return redirect('holiday_list')

    if request.method == 'POST':
        date_str = request.POST.get('date')
        name = request.POST.get('name', '').strip()
        if not date_str or not name:
            messages.error(request, 'Date and name are required.')
            return render(request, 'attendance_app/holiday_form.html')

        hol, created = Holiday.objects.get_or_create(
            date=date_str,
            defaults={'name': name, 'holiday_type': 'public'}
        )
        if not created:
            messages.warning(request, 'A holiday already exists on this date.')
        else:
            log_activity(request.user, 'create', 'Holiday', hol, f'Added holiday {name}', request)
            messages.success(request, f'Holiday {name} added.')
        return redirect('holiday_list')

    return render(request, 'attendance_app/holiday_form.html')


@login_required
def holiday_delete(request, pk):
    if not has_perm(request.user, 'holiday_add'):
        messages.error(request, 'Access denied.')
        return redirect('holiday_list')

    holiday = get_object_or_404(Holiday, pk=pk)
    if request.method == 'POST':
        name = holiday.name
        holiday.delete()
        log_activity(request.user, 'delete', 'Holiday', description=f'Deleted holiday {name}', request=request)
        messages.success(request, f'Holiday {name} deleted.')
    return redirect('holiday_list')


@login_required
def generate_sundays(request):
    if not has_perm(request.user, 'holiday_add'):
        messages.error(request, 'Access denied.')
        return redirect('holiday_list')

    year = int(request.POST.get('year', timezone.localdate().year))
    count = 0
    for month in range(1, 13):
        cal = calendar.monthcalendar(year, month)
        for week in cal:
            sunday = week[6]
            if sunday:
                date = datetime.date(year, month, sunday)
                _, created = Holiday.objects.get_or_create(
                    date=date,
                    defaults={'name': 'Sunday', 'holiday_type': 'sunday'}
                )
                if created:
                    count += 1

    log_activity(request.user, 'create', 'Holiday', description=f'Generated {count} Sundays for {year}', request=request)
    messages.success(request, f'{count} Sunday holidays generated for {year}.')
    return redirect('holiday_list')


# ─── LEAVES ──────────────────────────────────────────────────────────────────

@login_required
def leave_list(request):
    if not has_perm(request.user, 'leave_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    leaves = LeaveRequest.objects.select_related('employee', 'leave_type').all()
    status_filter = request.GET.get('status', '')
    if status_filter:
        leaves = leaves.filter(status=status_filter)
    return render(request, 'attendance_app/leave_list.html', {
        'leaves': leaves, 'status_filter': status_filter,
    })


@login_required
def leave_add(request):
    if not has_perm(request.user, 'leave_view'):
        messages.error(request, 'Access denied.')
        return redirect('leave_list')

    if request.method == 'POST':
        emp_id = request.POST.get('employee')
        lt_id = request.POST.get('leave_type')
        start = request.POST.get('start_date')
        end = request.POST.get('end_date')
        reason = request.POST.get('reason', '')

        lr = LeaveRequest.objects.create(
            employee_id=emp_id, leave_type_id=lt_id,
            start_date=start, end_date=end, reason=reason
        )
        log_activity(request.user, 'create', 'LeaveRequest', lr, f'Leave request for {lr.employee.name}', request)
        messages.success(request, 'Leave request submitted.')
        return redirect('leave_list')

    employees = Employee.objects.filter(status='active')
    leave_types = LeaveType.objects.all()
    return render(request, 'attendance_app/leave_form.html', {
        'employees': employees, 'leave_types': leave_types,
    })


@login_required
def leave_approve(request, pk):
    if not has_perm(request.user, 'leave_manage'):
        messages.error(request, 'Access denied.')
        return redirect('leave_list')

    lr = get_object_or_404(LeaveRequest, pk=pk)
    lr.status = 'approved'
    lr.reviewed_by = request.user
    lr.review_note = request.POST.get('note', '')
    lr.save()
    log_activity(request.user, 'update', 'LeaveRequest', lr, f'Approved leave for {lr.employee.name}', request)
    messages.success(request, 'Leave approved.')
    return redirect('leave_list')


@login_required
def leave_reject(request, pk):
    if not has_perm(request.user, 'leave_manage'):
        messages.error(request, 'Access denied.')
        return redirect('leave_list')

    lr = get_object_or_404(LeaveRequest, pk=pk)
    lr.status = 'rejected'
    lr.reviewed_by = request.user
    lr.review_note = request.POST.get('note', '')
    lr.save()
    log_activity(request.user, 'update', 'LeaveRequest', lr, f'Rejected leave for {lr.employee.name}', request)
    messages.success(request, 'Leave rejected.')
    return redirect('leave_list')


@login_required
def leave_type_list(request):
    leave_types = LeaveType.objects.all()
    return render(request, 'attendance_app/leave_type_list.html', {'leave_types': leave_types})


@login_required
def leave_type_add(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        days = request.POST.get('days_allowed', 0)
        desc = request.POST.get('description', '')
        if name:
            lt = LeaveType.objects.create(name=name, days_allowed=days, description=desc)
            log_activity(request.user, 'create', 'LeaveType', lt, f'Added leave type {name}', request)
            messages.success(request, f'Leave type {name} added.')
        return redirect('leave_type_list')
    return render(request, 'attendance_app/leave_type_form.html')


# ─── EXPORT ──────────────────────────────────────────────────────────────────

@login_required
def export_attendance(request):
    if not has_perm(request.user, 'export_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    today = timezone.localdate()
    return render(request, 'attendance_app/export.html', {
        'today': today,
        'months': [(i, datetime.date(2000, i, 1).strftime('%B')) for i in range(1, 13)],
        'years': range(2020, today.year + 2),
    })


@login_required
def export_get_employees(request):
    """AJAX: Return employees who attended in selected month"""
    year = int(request.GET.get('year', timezone.localdate().year))
    month = int(request.GET.get('month', timezone.localdate().month))

    employee_ids = Attendance.objects.filter(
        date__year=year, date__month=month,
        status__in=['present', 'half_day']
    ).values_list('employee_id', flat=True).distinct()

    employees = Employee.objects.filter(id__in=employee_ids).values('id', 'name')
    return JsonResponse({'employees': list(employees)})


@login_required
def export_pdf(request):
    if not has_perm(request.user, 'export_view'):
        return HttpResponse('Access denied', status=403)

    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    employee_ids = request.POST.getlist('employees')
    date_from_str = request.POST.get('date_from')
    date_to_str = request.POST.get('date_to')

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.graphics.shapes import Rect, Drawing
    except ImportError:
        return HttpResponse('ReportLab not installed. Run: pip install reportlab', status=500)

    date_from = datetime.date.fromisoformat(date_from_str) if date_from_str else datetime.date(year, month, 1)
    days_in_month = calendar.monthrange(year, month)[1]
    date_to = datetime.date.fromisoformat(date_to_str) if date_to_str else datetime.date(year, month, days_in_month)

    employees = list(Employee.objects.filter(id__in=employee_ids))
    holidays = {h.date: h.name for h in Holiday.objects.filter(date__gte=date_from, date__lte=date_to)}

    att_map = {}
    for att in Attendance.objects.filter(
        employee_id__in=employee_ids,
        date__gte=date_from, date__lte=date_to
    ).select_related('employee'):
        att_map[(att.employee_id, att.date)] = att

    cs = CompanySettings.get_settings()
    company_name    = cs.company_name
    company_address = cs.company_address
    company_phone   = cs.company_phone
    company_email   = cs.company_email
    logo_path       = cs.logo.path if cs.logo else None
    month_name      = datetime.date(year, month, 1).strftime('%B %Y')

    # ── COLORS ────────────────────────────────────────────────────────────────
    BLACK        = colors.black
    WHITE        = colors.white
    YELLOW       = colors.HexColor('#FFFF00')
    ABSENT_RED   = colors.HexColor('#CC0000')
    LEAVE_PURPLE = colors.HexColor('#7C3AED')
    HEADER_GRAY  = colors.HexColor('#E8E8E8')
    DARK_GRAY    = colors.HexColor('#404040')
    ORANGE       = colors.HexColor('#E8650A')
    LIGHT_ORANGE = colors.HexColor('#FFF3E8')

    # ── PAGE SETUP ────────────────────────────────────────────────────────────
    PAGE_W, PAGE_H = A4          # 595.27 x 841.89 pt
    L_MARGIN = R_MARGIN = 1.5 * cm
    T_MARGIN = B_MARGIN = 1.5 * cm

    # Usable width after margins
    usable_w = PAGE_W - L_MARGIN - R_MARGIN   # ~547 pt  ≈ 19.3 cm

    # 5 employees per page — compute exact col widths to fill full usable width
    EMPS_PER_PAGE = 5
    emp_chunks = [employees[i:i+EMPS_PER_PAGE] for i in range(0, len(employees), EMPS_PER_PAGE)]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=L_MARGIN, rightMargin=R_MARGIN,
        topMargin=T_MARGIN, bottomMargin=B_MARGIN
    )

    has_logo = bool(logo_path and os.path.exists(logo_path))

    story = []

    for chunk_idx, chunk in enumerate(emp_chunks):
        n = len(chunk)

        # DATE col + n*3 sub-cols — calculated to fill full page width exactly
        DATE_W = 1.75 * cm
        remaining = usable_w - DATE_W
        SUB_W = remaining / (n * 3)
        col_widths = [DATE_W] + [SUB_W] * (n * 3)

        # ── PROFESSIONAL HEADER ───────────────────────────────────────────
        header_w = sum(col_widths)

        co_name_style = ParagraphStyle('co', alignment=TA_CENTER, fontSize=16,
                                       fontName='Helvetica-Bold', leading=20,
                                       textColor=DARK_GRAY)
        co_addr_style = ParagraphStyle('addr', alignment=TA_CENTER, fontSize=8,
                                       leading=11, textColor=colors.HexColor('#666666'))
        co_sub_style  = ParagraphStyle('sub', alignment=TA_CENTER, fontSize=8,
                                       leading=11, textColor=colors.HexColor('#888888'))
        report_style  = ParagraphStyle('rep', alignment=TA_RIGHT, fontSize=9,
                                       fontName='Helvetica-Bold', leading=12, textColor=ORANGE)
        period_style  = ParagraphStyle('per', alignment=TA_RIGHT, fontSize=8,
                                       leading=11, textColor=DARK_GRAY)

        if has_logo:
            try:
                from reportlab.platypus import Image as RLImage
                logo_img = RLImage(logo_path, width=1.6*cm, height=1.6*cm)
                logo_cell = logo_img
            except Exception:
                logo_cell = ''
        else:
            # Placeholder box with "LOGO" text
            logo_cell = Paragraph(
                '<font size="7" color="#AAAAAA">LOGO</font>',
                ParagraphStyle('lp', alignment=TA_CENTER, fontSize=7,
                               textColor=colors.HexColor('#AAAAAA'))
            )

        contact_line = ''
        if company_phone:
            contact_line += f'Tel: {company_phone}'
        if company_email:
            if contact_line:
                contact_line += f'  |  Email: {company_email}'
            else:
                contact_line += f'Email: {company_email}'

        center_cell = [
            Paragraph(company_name, co_name_style),
            Paragraph(company_address, co_addr_style),
        ]
        if contact_line:
            center_cell.append(Paragraph(contact_line, co_sub_style))

        right_cell = [
            Paragraph('ATTENDANCE REPORT', report_style),
            Paragraph(month_name, period_style),
            Paragraph(
                f'{date_from.strftime("%d %b %Y")} — {date_to.strftime("%d %b %Y")}',
                ParagraphStyle('pd', alignment=TA_RIGHT, fontSize=7, leading=10,
                               textColor=colors.HexColor('#888888'))
            ),
        ]

        logo_w = 1.8 * cm
        right_w = 4.5 * cm
        center_w = header_w - logo_w - right_w

        header_tbl = Table(
            [[logo_cell, center_cell, right_cell]],
            colWidths=[logo_w, center_w, right_w]
        )
        header_tbl.setStyle(TableStyle([
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
            ('ALIGN',         (0,0),(0,0),   'CENTER'),
            ('ALIGN',         (1,0),(1,0),   'CENTER'),
            ('ALIGN',         (2,0),(2,0),   'RIGHT'),
            ('LEFTPADDING',   (0,0),(-1,-1), 4),
            ('RIGHTPADDING',  (0,0),(-1,-1), 4),
            ('TOPPADDING',    (0,0),(-1,-1), 3),
            ('BOTTOMPADDING', (0,0),(-1,-1), 3),
            # Orange left border on logo cell as accent
            ('LINEAFTER',     (0,0),(0,0),   0.5, colors.HexColor('#DDDDDD')),
            # Logo cell light bg
            ('BACKGROUND',    (0,0),(0,0),   colors.HexColor('#FFF8F3')),
            ('BOX',           (0,0),(-1,-1), 1.5, ORANGE),
            # Light background for entire header
            ('BACKGROUND',    (0,0),(-1,-1), LIGHT_ORANGE),
        ]))

        story.append(header_tbl)

        # Orange separator line
        story.append(HRFlowable(
            width=header_w, thickness=2.5, color=ORANGE,
            spaceAfter=3, spaceBefore=1
        ))

        # ── NAME ROW (row 0) ──────────────────────────────────────────────
        name_row = ['NAME']
        for emp in chunk:
            name_row.extend([emp.name, '', ''])

        # ── DATE/IN/OUT/OT SUB-HEADER (row 1) ────────────────────────────
        sub_row = ['DATE']
        for _ in chunk:
            sub_row.extend(['IN', 'OUT', 'OT'])

        # ── DATA ROWS ─────────────────────────────────────────────────────
        data_rows = []
        holiday_row_indices = []
        date_list = []

        current = date_from
        while current <= date_to:
            is_holiday = current in holidays or current.weekday() == 6
            row = [current.strftime('%d-%m-%Y')]

            for emp in chunk:
                att = att_map.get((emp.id, current))
                if att:
                    if att.status == 'absent':
                        row.extend(['A', '', ''])
                    elif att.status == 'leave':
                        row.extend(['V', '', ''])
                    elif att.status == 'holiday':
                        row.extend(['', '', ''])
                    elif att.status == 'half_day':
                        in_t  = att.in_time.strftime('%H:%M')  if att.in_time  else ''
                        out_t = att.out_time.strftime('%H:%M') if att.out_time else ''
                        ot    = str(att.ot_hours) if att.ot_hours else ''
                        row.extend([in_t, out_t, ot])
                    else:  # present
                        in_t  = att.in_time.strftime('%H:%M')  if att.in_time  else ''
                        out_t = att.out_time.strftime('%H:%M') if att.out_time else ''
                        ot    = str(att.ot_hours) if att.ot_hours else ''
                        row.extend([in_t, out_t, ot])
                else:
                    row.extend(['', '', ''] if is_holiday else ['A', '', ''])

            date_list.append(current)
            if is_holiday:
                holiday_row_indices.append(len(data_rows))
            data_rows.append(row)
            current += datetime.timedelta(days=1)

        # ── TOTALS ────────────────────────────────────────────────────────
        total_ot_row    = ['Total OT']
        holiday_ot_row  = ['Holiday OT']
        absent_cnt_row  = ['Absent']

        for emp in chunk:
            # Holiday OT = ot_hours on holiday/sunday days ONLY
            hol_ot = sum(
                float(att_map[(emp.id, d)].ot_hours)
                for d in date_list
                if (d in holidays or d.weekday() == 6)
                and (emp.id, d) in att_map
                and att_map[(emp.id, d)].ot_hours
            )
            # Normal OT = ot_hours on regular working days ONLY (excludes holiday/sunday)
            total_ot = sum(
                float(att_map[(emp.id, d)].ot_hours)
                for d in date_list
                if d not in holidays and d.weekday() != 6
                and (emp.id, d) in att_map
                and att_map[(emp.id, d)].ot_hours
            )
            # Count BOTH absent and leave days
            absent_cnt = sum(
                1 for d in date_list
                if (emp.id, d) in att_map and att_map[(emp.id, d)].status == 'absent'
            )
            leave_cnt = sum(
                1 for d in date_list
                if (emp.id, d) in att_map and att_map[(emp.id, d)].status == 'leave'
            )
            # Also count days with no record and no holiday = absent
            for d in date_list:
                if (emp.id, d) not in att_map and d not in holidays and d.weekday() != 6:
                    absent_cnt += 1

            absent_leave_fmt = f'A:{absent_cnt} / L:{leave_cnt}' if (absent_cnt or leave_cnt) else ''
            total_ot_row.extend([f'{total_ot:.1f}' if total_ot else '', '', ''])
            holiday_ot_row.extend([f'{hol_ot:.1f}' if hol_ot else '', '', ''])
            absent_cnt_row.extend([absent_leave_fmt, '', ''])

        # ── BUILD FULL TABLE ──────────────────────────────────────────────
        all_rows   = [name_row, sub_row] + data_rows + [total_ot_row, holiday_ot_row, absent_cnt_row]
        data_start = 2
        total_start = data_start + len(data_rows)

        # Row heights: header rows taller, data rows compact to fill page
        # Estimate available height for data rows
        header_h  = 2.0 * cm   # header block
        sep_h     = 0.35 * cm
        name_row_h = 0.9 * cm
        sub_row_h  = 1 * cm
        totals_h   = 0.62 * cm * 3
        available  = PAGE_H - T_MARGIN - B_MARGIN - header_h - sep_h - name_row_h - sub_row_h - totals_h
        row_h = max(0.65 * cm, min(0.55 * cm, available / max(len(data_rows), 1)))

        row_heights = [name_row_h, sub_row_h] + [row_h] * len(data_rows) + [0.62*cm, 0.62*cm, 0.62*cm]

        tbl = Table(all_rows, colWidths=col_widths, rowHeights=row_heights, repeatRows=2)

        # ── STYLES ────────────────────────────────────────────────────────
        s = [
            ('GRID',          (0,0), (-1,-1), 0.5, BLACK),
            ('FONTSIZE',      (0,0), (-1,-1), 7),
            ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING',    (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
            ('LEFTPADDING',   (0,0), (-1,-1), 1),
            ('RIGHTPADDING',  (0,0), (-1,-1), 1),

            # NAME row
            ('BACKGROUND',    (0,0), (-1,0), HEADER_GRAY),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,0), 7),

            # Sub-header row
            ('BACKGROUND',    (0,1), (-1,1), HEADER_GRAY),
            ('FONTNAME',      (0,1), (-1,1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,1), (-1,1), 7),

            # DATE column
            ('FONTNAME',      (0,2), (0,-1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,2), (0,-4), 7),

            # Totals rows
            ('BACKGROUND',    (0,total_start), (-1,-1), HEADER_GRAY),
            ('FONTNAME',      (0,total_start), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,total_start), (-1,-1), 7),
            ('ALIGN',         (0,total_start), (0,-1),  'LEFT'),
            ('LEFTPADDING',   (0,total_start), (0,-1),  3),
        ]

        # Span employee names in NAME row
        for i in range(n):
            cs = 1 + i * 3
            ce = cs + 2
            s.append(('SPAN',     (cs, 0), (ce, 0)))
            s.append(('FONTSIZE', (cs, 0), (ce, 0), 7))

        # Span totals values across 3 cols per employee
        for tr in range(3):
            ri = total_start + tr
            for i in range(n):
                cs = 1 + i * 3
                ce = cs + 2
                s.append(('SPAN', (cs, ri), (ce, ri)))

        # Yellow holiday rows — full row
        for ri in holiday_row_indices:
            s.append(('BACKGROUND', (0, ri + data_start), (-1, ri + data_start), YELLOW))

        # ABSENT = red bold, LEAVE = purple bold (per cell)
        for ri, row_data in enumerate(data_rows):
            ari = ri + data_start
            for ci, cell in enumerate(row_data):
                v = str(cell)
                if v == 'ABSENT':
                    s.append(('BACKGROUND', (ci, ari), (ci, ari), ABSENT_RED))
                    s.append(('FONTNAME',  (ci, ari), (ci, ari), 'Helvetica-Bold'))
                elif v == 'LEAVE':
                    s.append(('BACKGROUND', (ci, ari), (ci, ari), LEAVE_PURPLE))
                    s.append(('FONTNAME',  (ci, ari), (ci, ari), 'Helvetica-Bold'))

        tbl.setStyle(TableStyle(s))
        story.append(tbl)

        if chunk_idx < len(emp_chunks) - 1:
            story.append(PageBreak())

    doc.build(story)
    buffer.seek(0)

    log_activity(request.user, 'export', 'Attendance', description=f'PDF export {month_name}', request=request)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attendance_{year}_{month:02d}.pdf"'
    return response


@login_required
def export_excel(request):
    if not has_perm(request.user, 'export_view'):
        return HttpResponse('Access denied', status=403)

    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    employee_ids = request.POST.getlist('employees')
    date_from_str = request.POST.get('date_from')
    date_to_str = request.POST.get('date_to')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed. Run: pip install openpyxl', status=500)

    date_from = datetime.date.fromisoformat(date_from_str) if date_from_str else datetime.date(year, month, 1)
    days_in_month = calendar.monthrange(year, month)[1]
    date_to = datetime.date.fromisoformat(date_to_str) if date_to_str else datetime.date(year, month, days_in_month)
    employees = list(Employee.objects.filter(id__in=employee_ids))
    holidays = {h.date: h.name for h in Holiday.objects.filter(date__gte=date_from, date__lte=date_to)}

    # Pre-fetch all attendance
    att_map = {}
    for att in Attendance.objects.filter(
        employee_id__in=employee_ids,
        date__gte=date_from, date__lte=date_to
    ).select_related('employee'):
        att_map[(att.employee_id, att.date)] = att

    cs = CompanySettings.get_settings()
    company_name = cs.company_name
    company_address = cs.company_address
    month_name = datetime.date(year, month, 1).strftime('%B %Y')

    # Styles
    YELLOW_FILL  = PatternFill("solid", fgColor="FFFF00")
    HEADER_FILL  = PatternFill("solid", fgColor="D9D9D9")
    WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
    ABSENT_FILL  = PatternFill("solid", fgColor="FFFFFF")

    bold_font    = Font(bold=True, size=8)
    normal_font  = Font(size=7)
    absent_font  = Font(bold=True, size=7, color="FF0000")
    leave_font   = Font(bold=True, size=7, color="8B5CF6")
    title_font   = Font(bold=True, size=14)
    addr_font    = Font(size=9, color="444444")
    center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_center  = Alignment(horizontal='left', vertical='center')

    def thin_border():
        s = Side(style='thin', color='000000')
        return Border(left=s, right=s, top=s, bottom=s)

    BORDER = thin_border()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance {month_name}"

    # 5 employees per sheet (can add more sheets if needed)
    EMPS_PER_PAGE = 5
    emp_chunks = [employees[i:i+EMPS_PER_PAGE] for i in range(0, len(employees), EMPS_PER_PAGE)]

    current_row = 1

    for chunk_idx, chunk in enumerate(emp_chunks):
        n = len(chunk)
        # Total columns: 1 (DATE) + n*3
        total_cols = 1 + n * 3

        if chunk_idx > 0:
            current_row += 2  # spacer between chunks

        # ── COMPANY HEADER ──────────────────────────────────────────────
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        c = ws.cell(row=current_row, column=1, value=company_name)
        c.font = title_font
        c.alignment = center
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        c = ws.cell(row=current_row, column=1, value=company_address)
        c.font = addr_font
        c.alignment = center
        ws.row_dimensions[current_row].height = 14
        current_row += 1

        # ── NAME ROW ─────────────────────────────────────────────────────
        name_row_idx = current_row
        ws.cell(row=current_row, column=1, value='NAME').font = bold_font
        ws.cell(row=current_row, column=1).fill = HEADER_FILL
        ws.cell(row=current_row, column=1).alignment = center
        ws.cell(row=current_row, column=1).border = BORDER

        for i, emp in enumerate(chunk):
            c_start = 2 + i * 3
            c_end   = c_start + 2
            ws.merge_cells(start_row=current_row, start_column=c_start, end_row=current_row, end_column=c_end)
            c = ws.cell(row=current_row, column=c_start, value=emp.name)
            c.font = bold_font
            c.fill = HEADER_FILL
            c.alignment = center
            c.border = BORDER
            # borders on merged cells
            for col in range(c_start, c_end + 1):
                ws.cell(row=current_row, column=col).border = BORDER

        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # ── DATE/IN/OUT/OT SUB-HEADER ROW ────────────────────────────────
        sub_row_idx = current_row
        for col_i, val in enumerate(['DATE'] + ['IN', 'OUT', 'OT'] * n, 1):
            c = ws.cell(row=current_row, column=col_i, value=val)
            c.font = bold_font
            c.fill = HEADER_FILL
            c.alignment = center
            c.border = BORDER
        ws.row_dimensions[current_row].height = 14
        current_row += 1

        # ── DATA ROWS ─────────────────────────────────────────────────────
        date_rows_start = current_row
        date_list = []
        cur = date_from
        while cur <= date_to:
            date_list.append(cur)
            cur += datetime.timedelta(days=1)

        for cur in date_list:
            is_hol = cur in holidays or cur.weekday() == 6
            row_fill = YELLOW_FILL if is_hol else WHITE_FILL

            # DATE cell
            dc = ws.cell(row=current_row, column=1, value=cur.strftime('%d-%m-%Y'))
            dc.font = Font(bold=True, size=7)
            dc.fill = row_fill
            dc.alignment = center
            dc.border = BORDER

            for i, emp in enumerate(chunk):
                att = att_map.get((emp.id, cur))
                c_base = 2 + i * 3

                if att:
                    if att.status == 'absent':
                        vals = ['ABSENT', '', '']
                    elif att.status == 'leave':
                        vals = ['LEAVE', '', '']
                    elif att.status == 'holiday':
                        vals = ['', '', '']
                    elif att.status == 'half_day':
                        vals = [
                            att.in_time.strftime('%H:%M') if att.in_time else '',
                            att.out_time.strftime('%H:%M') if att.out_time else '',
                            str(att.ot_hours) if att.ot_hours else ''
                        ]
                    else:
                        vals = [
                            att.in_time.strftime('%H:%M') if att.in_time else '',
                            att.out_time.strftime('%H:%M') if att.out_time else '',
                            str(att.ot_hours) if att.ot_hours else ''
                        ]
                else:
                    vals = ['', '', ''] if is_hol else ['ABSENT', '', '']

                for j, v in enumerate(vals):
                    c = ws.cell(row=current_row, column=c_base + j, value=v)
                    c.fill = row_fill
                    c.alignment = center
                    c.border = BORDER
                    if v == 'ABSENT':
                        c.font = absent_font
                    elif v == 'LEAVE':
                        c.font = leave_font
                    else:
                        c.font = normal_font

            ws.row_dimensions[current_row].height = 13
            current_row += 1

        # ── TOTALS ROWS ───────────────────────────────────────────────────
        totals_labels = ['Total OT', 'Holiday OT', 'Absent / Leave']
        totals_values = [[], [], []]

        for emp in chunk:
            # Holiday OT = ot_hours on holiday/sunday days ONLY
            hol_ot = sum(
                float(att_map[(emp.id, d)].ot_hours)
                for d in date_list
                if (d in holidays or d.weekday() == 6)
                and (emp.id, d) in att_map
                and att_map[(emp.id, d)].ot_hours
            )
            # Normal OT = ot_hours on regular working days ONLY (excludes holiday/sunday)
            total_ot = sum(
                float(att_map[(emp.id, d)].ot_hours)
                for d in date_list
                if d not in holidays and d.weekday() != 6
                and (emp.id, d) in att_map
                and att_map[(emp.id, d)].ot_hours
            )
            absent_cnt = sum(
                1 for d in date_list
                if (emp.id, d) in att_map and att_map[(emp.id, d)].status == 'absent'
            )
            leave_cnt = sum(
                1 for d in date_list
                if (emp.id, d) in att_map and att_map[(emp.id, d)].status == 'leave'
            )
            # Days with no record and not a holiday = absent
            for d in date_list:
                if (emp.id, d) not in att_map and d not in holidays and d.weekday() != 6:
                    absent_cnt += 1

            absent_leave = f'A:{absent_cnt} / L:{leave_cnt}' if (absent_cnt or leave_cnt) else ''
            totals_values[0].append(f'{total_ot:.1f}' if total_ot else '')
            totals_values[1].append(f'{hol_ot:.1f}'   if hol_ot  else '')
            totals_values[2].append(absent_leave)

        for t_idx, label in enumerate(totals_labels):
            # Label cell
            lc = ws.cell(row=current_row, column=1, value=label)
            lc.font = bold_font
            lc.fill = HEADER_FILL
            lc.alignment = left_center
            lc.border = BORDER

            for i in range(n):
                c_start = 2 + i * 3
                c_end   = c_start + 2
                ws.merge_cells(start_row=current_row, start_column=c_start, end_row=current_row, end_column=c_end)
                vc = ws.cell(row=current_row, column=c_start, value=totals_values[t_idx][i])
                vc.font = bold_font
                vc.fill = HEADER_FILL
                vc.alignment = center
                for col in range(c_start, c_end + 1):
                    ws.cell(row=current_row, column=col).border = BORDER

            ws.row_dimensions[current_row].height = 14
            current_row += 1

        # ── COLUMN WIDTHS ─────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 13
        for i in range(n):
            base = 2 + i * 3
            ws.column_dimensions[get_column_letter(base)].width = 9
            ws.column_dimensions[get_column_letter(base + 1)].width = 9
            ws.column_dimensions[get_column_letter(base + 2)].width = 6

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_activity(request.user, 'export', 'Attendance', description=f'Excel export {month_name}', request=request)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_{year}_{month:02d}.xlsx"'
    return response

    wb = openpyxl.Workbook()
    ws = wb.active
    company_name = getattr(settings, 'COMPANY_NAME', 'Your Company')
    month_name = datetime.date(year, month, 1).strftime('%B %Y')
    ws.title = f"Attendance {month_name}"

    orange_fill = PatternFill("solid", fgColor="F97316")
    header_fill = PatternFill("solid", fgColor="1E293B")
    absent_fill = PatternFill("solid", fgColor="FEE2E2")
    holiday_fill = PatternFill("solid", fgColor="EDE9FE")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    bold_white = Font(bold=True, color="FFFFFF")
    bold_orange = Font(bold=True, color="F97316", size=14)
    absent_font = Font(bold=True, color="EF4444")
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = company_name
    ws['A1'].font = bold_orange
    ws.row_dimensions[1].height = 25

    ws.merge_cells('C1:E1')
    ws['C1'] = f'Attendance Report - {month_name}'
    ws['C1'].font = Font(bold=True, size=12)
    ws['C1'].alignment = center

    ws.append([])

    emp_list = list(employees)
    # Header row
    header = ['Date']
    for emp in emp_list:
        header.extend([emp.name, 'OUT', 'OT'])

    ws.append(header)
    header_row = ws.max_row
    for col_idx, val in enumerate(header, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = orange_fill if col_idx == 1 else header_fill
        cell.font = bold_white
        cell.alignment = center
        cell.border = thin

    # Merge employee name headers
    for emp_idx in range(len(emp_list)):
        start_col = 2 + emp_idx * 3
        if start_col < start_col + 1:
            pass

    ws.column_dimensions['A'].width = 14
    for emp_idx in range(len(emp_list)):
        base = 2 + emp_idx * 3
        ws.column_dimensions[get_column_letter(base)].width = 12
        ws.column_dimensions[get_column_letter(base + 1)].width = 10
        ws.column_dimensions[get_column_letter(base + 2)].width = 8

    current = date_from
    row_num = header_row + 1
    while current <= date_to:
        row_data = [current.strftime('%d %b %Y (%a)')]
        is_holiday = current in holidays
        is_sunday = current.weekday() == 6

        for emp in emp_list:
            try:
                att = Attendance.objects.get(employee=emp, date=current)
                if att.status == 'absent':
                    row_data.extend(['ABSENT', '', ''])
                elif att.status == 'half_day':
                    in_t = att.in_time.strftime('%H:%M') if att.in_time else '-'
                    out_t = att.out_time.strftime('%H:%M') if att.out_time else '-'
                    row_data.extend([f'HD {in_t}', out_t, str(att.ot_hours)])
                else:
                    in_t = att.in_time.strftime('%H:%M') if att.in_time else '-'
                    out_t = att.out_time.strftime('%H:%M') if att.out_time else '-'
                    row_data.extend([in_t, out_t, str(att.ot_hours) if att.ot_hours else ''])
            except Attendance.DoesNotExist:
                if is_holiday or is_sunday:
                    hname = holidays.get(current, 'Sunday' if is_sunday else 'Holiday')
                    row_data.extend([hname, '', ''])
                else:
                    row_data.extend(['ABSENT', '', ''])

        ws.append(row_data)
        fill = alt_fill if row_num % 2 == 0 else white_fill
        if is_holiday or is_sunday:
            fill = holiday_fill

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = fill
            cell.alignment = center
            cell.border = thin
            if str(val) == 'ABSENT':
                cell.fill = absent_fill
                cell.font = absent_font

        row_num += 1
        current += datetime.timedelta(days=1)




# ─── SETTINGS ─────────────────────────────────────────────────────────────────

@login_required
def settings_view(request):
    if not request.user.is_superuser:
        messages.error(request, 'Only administrators can access settings.')
        return redirect('dashboard')

    cs = CompanySettings.get_settings()

    if request.method == 'POST':
        tab = request.POST.get('tab', 'company')

        if tab == 'company':
            cs.company_name    = request.POST.get('company_name', '').strip() or cs.company_name
            cs.company_address = request.POST.get('company_address', '').strip()
            cs.company_phone   = request.POST.get('company_phone', '').strip()
            cs.company_email   = request.POST.get('company_email', '').strip()
            cs.company_website = request.POST.get('company_website', '').strip()
            cs.company_trn     = request.POST.get('company_trn', '').strip()
            if request.FILES.get('logo'):
                cs.logo = request.FILES['logo']
            if request.POST.get('remove_logo') == '1':
                cs.logo = None
            cs.save()
            # Also update settings.py runtime values
            from django.conf import settings as django_settings
            django_settings.COMPANY_NAME    = cs.company_name
            django_settings.COMPANY_ADDRESS = cs.company_address
            django_settings.COMPANY_PHONE   = cs.company_phone
            django_settings.COMPANY_EMAIL   = cs.company_email
            log_activity(request.user, 'update', 'Settings', description='Updated company profile', request=request)
            messages.success(request, 'Company profile saved successfully.')

        elif tab == 'work':
            in_str  = request.POST.get('default_in_time', '')
            out_str = request.POST.get('default_out_time', '')
            if in_str:
                cs.default_in_time = datetime.time.fromisoformat(in_str)
            if out_str:
                cs.default_out_time = datetime.time.fromisoformat(out_str)
            cs.work_days   = request.POST.get('work_days', cs.work_days)
            cs.weekend_day = request.POST.get('weekend_day', cs.weekend_day)
            cs.save()
            log_activity(request.user, 'update', 'Settings', description='Updated work schedule', request=request)
            messages.success(request, 'Work schedule saved successfully.')

        elif tab == 'password':
            current_pw  = request.POST.get('current_password', '')
            new_pw      = request.POST.get('new_password', '')
            confirm_pw  = request.POST.get('confirm_password', '')
            if not request.user.check_password(current_pw):
                messages.error(request, 'Current password is incorrect.')
            elif len(new_pw) < 6:
                messages.error(request, 'New password must be at least 6 characters.')
            elif new_pw != confirm_pw:
                messages.error(request, 'New passwords do not match.')
            else:
                request.user.set_password(new_pw)
                request.user.save()
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, request.user)
                log_activity(request.user, 'update', 'Auth', description='Changed admin password', request=request)
                messages.success(request, 'Password changed successfully.')

        elif tab == 'system':
            cs.timezone_name = request.POST.get('timezone_name', cs.timezone_name)
            cs.date_format   = request.POST.get('date_format', cs.date_format)
            cs.save()
            log_activity(request.user, 'update', 'Settings', description='Updated system settings', request=request)
            messages.success(request, 'System settings saved.')

        return redirect(f"/settings/?tab={tab}")

    active_tab = request.GET.get('tab', 'company')
    timezones = [
        'Asia/Dubai','Asia/Riyadh','Asia/Kuwait','Asia/Qatar','Asia/Bahrain',
        'Asia/Muscat','Asia/Karachi','Asia/Kolkata','Asia/Dhaka','Asia/Bangkok',
        'Asia/Singapore','Asia/Manila','Asia/Jakarta','Africa/Cairo','Africa/Nairobi',
        'Europe/London','Europe/Paris','Europe/Istanbul','America/New_York','America/Chicago',
        'America/Denver','America/Los_Angeles','Australia/Sydney','Pacific/Auckland',
    ]
    return render(request, 'attendance_app/settings.html', {
        'cs': cs,
        'active_tab': active_tab,
        'timezones': timezones,
    })

# ─── USER MANAGEMENT ─────────────────────────────────────────────────────────

@login_required
def user_list(request):
    if not request.user.is_superuser and not has_perm(request.user, 'user_manage'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    users = User.objects.all().select_related('custom_permissions')
    return render(request, 'attendance_app/user_list.html', {'users': users})


@login_required
def user_add(request):
    if not request.user.is_superuser and not has_perm(request.user, 'user_manage'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    all_perms = UserPermission.PERMISSION_CHOICES

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        email = request.POST.get('email', '')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        selected_perms = request.POST.getlist('permissions')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return render(request, 'attendance_app/user_form.html', {'all_perms': all_perms, 'action': 'Add'})

        user = User.objects.create_user(
            username=username, password=password,
            email=email, first_name=first_name, last_name=last_name
        )
        UserPermission.objects.create(user=user, permissions=selected_perms)
        log_activity(request.user, 'create', 'User', user, f'Created user {username}', request)
        messages.success(request, f'User {username} created.')
        return redirect('user_list')

    return render(request, 'attendance_app/user_form.html', {'all_perms': all_perms, 'action': 'Add'})


@login_required
def user_edit(request, pk):
    if not request.user.is_superuser and not has_perm(request.user, 'user_manage'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    user_obj = get_object_or_404(User, pk=pk)
    try:
        user_perm = user_obj.custom_permissions
    except UserPermission.DoesNotExist:
        user_perm = UserPermission.objects.create(user=user_obj, permissions=[])

    all_perms = UserPermission.PERMISSION_CHOICES

    if request.method == 'POST':
        user_obj.email = request.POST.get('email', '')
        user_obj.first_name = request.POST.get('first_name', '')
        user_obj.last_name = request.POST.get('last_name', '')
        new_password = request.POST.get('password', '')
        if new_password:
            user_obj.set_password(new_password)
        user_obj.save()
        user_perm.permissions = request.POST.getlist('permissions')
        user_perm.save()
        # Link/unlink employee portal
        emp_link_id = request.POST.get('employee_link', '')
        # Unlink current employee if any
        Employee.objects.filter(portal_user=user_obj).update(portal_user=None)
        if emp_link_id:
            try:
                emp_to_link = Employee.objects.get(pk=emp_link_id)
                emp_to_link.portal_user = user_obj
                emp_to_link.save()
            except Employee.DoesNotExist:
                pass
        log_activity(request.user, 'update', 'User', user_obj, f'Updated user {user_obj.username}', request)
        messages.success(request, f'User {user_obj.username} updated.')
        return redirect('user_list')

    # Employees not yet linked to any portal user (plus current one)
    linked_emp = None
    try:
        linked_emp = user_obj.employee_profile
    except Exception:
        pass
    unlinked_emps = Employee.objects.filter(status='active', portal_user__isnull=True).order_by('name')

    return render(request, 'attendance_app/user_form.html', {
        'user_obj': user_obj, 'user_perm': user_perm,
        'all_perms': all_perms, 'action': 'Edit',
        'unlinked_emps': unlinked_emps,
        'linked_emp': linked_emp,
    })


@login_required
def user_delete(request, pk):
    if not request.user.is_superuser:
        messages.error(request, 'Access denied.')
        return redirect('user_list')

    user_obj = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        if user_obj == request.user:
            messages.error(request, 'Cannot delete your own account.')
            return redirect('user_list')
        name = user_obj.username
        user_obj.delete()
        log_activity(request.user, 'delete', 'User', description=f'Deleted user {name}', request=request)
        messages.success(request, f'User {name} deleted.')
    return redirect('user_list')


# ─── ACTIVITY LOG ─────────────────────────────────────────────────────────────

@login_required
def activity_log(request):
    if not has_perm(request.user, 'activity_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    logs = ActivityLog.objects.select_related('user').all()
    action_filter = request.GET.get('action', '')
    model_filter = request.GET.get('model', '')
    user_filter = request.GET.get('user', '')
    date_filter = request.GET.get('date', '')

    if action_filter:
        logs = logs.filter(action=action_filter)
    if model_filter:
        logs = logs.filter(model_name__icontains=model_filter)
    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)
    if date_filter:
        logs = logs.filter(timestamp__date=date_filter)

    return render(request, 'attendance_app/activity_log.html', {
        'logs': logs[:200],
        'action_choices': ActivityLog.ACTION_CHOICES,
        'action_filter': action_filter,
        'model_filter': model_filter,
        'user_filter': user_filter,
        'date_filter': date_filter,
    })


# ─── EMPLOYEE SELF-SERVICE PORTAL ────────────────────────────────────────────

def portal_login_required(view_func):
    """Redirect portal users away from admin views; ensure login."""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        # Make sure this user has an employee profile
        try:
            emp = request.user.employee_profile
        except Exception:
            messages.error(request, 'No employee profile linked to your account.')
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper


@portal_login_required
def portal_dashboard(request):
    emp = request.user.employee_profile
    today = timezone.localdate()

    # This month stats
    month_start = today.replace(day=1)
    month_end_day = calendar.monthrange(today.year, today.month)[1]
    month_end = today.replace(day=month_end_day)

    holidays = {h.date for h in Holiday.objects.filter(date__gte=month_start, date__lte=month_end)}
    working_days = sum(
        1 for d in (month_start + datetime.timedelta(i) for i in range((month_end - month_start).days + 1))
        if d not in holidays and d.weekday() != 6 and d <= today
    )

    month_atts = list(Attendance.objects.filter(employee=emp, date__gte=month_start, date__lte=today))
    present  = sum(1 for a in month_atts if a.status == 'present')
    half_day = sum(1 for a in month_atts if a.status == 'half_day')
    absent   = sum(1 for a in month_atts if a.status == 'absent')
    leave    = sum(1 for a in month_atts if a.status == 'leave')
    total_ot = sum(float(a.ot_hours) for a in month_atts)
    total_hrs = sum(float(a.total_hours()) for a in month_atts)
    unmarked = max(0, working_days - present - half_day - absent - leave)
    att_rate = round((present + half_day * 0.5) / working_days * 100) if working_days else 0

    # Today's record
    try:
        today_att = Attendance.objects.get(employee=emp, date=today)
    except Attendance.DoesNotExist:
        today_att = None

    # Upcoming holidays
    upcoming_holidays = Holiday.objects.filter(date__gte=today).order_by('date')[:5]

    # Recent 7 records
    recent_atts = Attendance.objects.filter(employee=emp).order_by('-date')[:7]

    # 30-day trend for mini chart
    trend_labels, trend_status = [], []
    for i in range(29, -1, -1):
        d = today - datetime.timedelta(days=i)
        trend_labels.append(d.strftime('%d %b'))
        try:
            a = Attendance.objects.get(employee=emp, date=d)
            trend_status.append(a.status)
        except Attendance.DoesNotExist:
            if d in holidays or d.weekday() == 6:
                trend_status.append('holiday')
            else:
                trend_status.append('unmarked')

    return render(request, 'attendance_app/portal_dashboard.html', {
        'emp': emp,
        'today': today,
        'today_att': today_att,
        'present': present, 'half_day': half_day,
        'absent': absent + unmarked, 'leave': leave,
        'total_ot': round(total_ot, 1),
        'total_hrs': round(total_hrs, 1),
        'att_rate': att_rate,
        'working_days': working_days,
        'upcoming_holidays': upcoming_holidays,
        'recent_atts': recent_atts,
        'trend_labels': trend_labels,
        'trend_status': trend_status,
        'month_name': today.strftime('%B %Y'),
    })


@portal_login_required
def portal_attendance(request):
    emp = request.user.employee_profile
    today = timezone.localdate()

    year  = int(request.GET.get('year',  today.year))
    month = int(request.GET.get('month', today.month))

    date_from = datetime.date(year, month, 1)
    date_to   = datetime.date(year, month, calendar.monthrange(year, month)[1])

    holidays_qs = Holiday.objects.filter(date__gte=date_from, date__lte=date_to)
    holiday_map = {h.date: h.name for h in holidays_qs}
    holidays_set = set(holiday_map.keys())

    working_days = sum(
        1 for d in (date_from + datetime.timedelta(i) for i in range((date_to - date_from).days + 1))
        if d not in holidays_set and d.weekday() != 6
    )

    atts = {a.date: a for a in Attendance.objects.filter(employee=emp, date__gte=date_from, date__lte=date_to)}

    # Build calendar rows
    cal_days = []
    current = date_from
    while current <= date_to:
        is_holiday = current in holidays_set or current.weekday() == 6
        att = atts.get(current)
        holiday_name = holiday_map.get(current, 'Sunday' if current.weekday() == 6 else '')
        cal_days.append({
            'date': current,
            'att': att,
            'is_holiday': is_holiday,
            'holiday_name': holiday_name,
            'is_future': current > today,
            'is_today': current == today,
            'weekday': current.strftime('%a'),
        })
        current += datetime.timedelta(days=1)

    present  = sum(1 for a in atts.values() if a.status == 'present')
    half_day = sum(1 for a in atts.values() if a.status == 'half_day')
    absent   = sum(1 for a in atts.values() if a.status == 'absent')
    leave    = sum(1 for a in atts.values() if a.status == 'leave')
    total_ot = round(sum(float(a.ot_hours) for a in atts.values()), 1)
    unmarked = max(0, working_days - present - half_day - absent - leave)
    att_rate = round((present + half_day * 0.5) / working_days * 100) if working_days else 0

    return render(request, 'attendance_app/portal_attendance.html', {
        'emp': emp,
        'cal_days': cal_days,
        'year': year, 'month': month,
        'month_name': date_from.strftime('%B %Y'),
        'date_from': date_from, 'date_to': date_to,
        'present': present, 'half_day': half_day,
        'absent': absent + unmarked, 'leave': leave,
        'total_ot': total_ot,
        'working_days': working_days,
        'att_rate': att_rate,
        'months': [(i, datetime.date(2000, i, 1).strftime('%B')) for i in range(1, 13)],
        'years': range(2020, today.year + 2),
        'today': today,
    })


@login_required
def portal_attendance_ajax(request):
    """Return single day attendance detail as JSON for modal."""
    emp = getattr(request.user, 'employee_profile', None)
    if not emp:
        return JsonResponse({'error': 'No employee profile'}, status=403)
    date_str = request.GET.get('date', '')
    try:
        d = datetime.date.fromisoformat(date_str)
        att = Attendance.objects.get(employee=emp, date=d)
        return JsonResponse({
            'date': d.strftime('%d %B %Y'),
            'status': att.get_status_display(),
            'in_time': att.in_time.strftime('%H:%M') if att.in_time else '—',
            'out_time': att.out_time.strftime('%H:%M') if att.out_time else '—',
            'ot_hours': str(att.ot_hours),
            'total_hours': att.total_hours(),
            'notes': att.notes or '',
        })
    except (ValueError, Attendance.DoesNotExist):
        return JsonResponse({'error': 'No record'}, status=404)


# ─── REPORTS & ANALYTICS ──────────────────────────────────────────────────────

def _get_report_filters(request):
    """Parse common year/month/employee filters from GET params."""
    today = timezone.localdate()
    year  = int(request.GET.get('year',  today.year))
    month = int(request.GET.get('month', today.month))
    emp_id = request.GET.get('employee', '')
    date_from = datetime.date(year, month, 1)
    date_to   = datetime.date(year, month, calendar.monthrange(year, month)[1])
    return year, month, emp_id, date_from, date_to


@login_required
def reports_view(request):
    if not has_perm(request.user, 'export_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    today = timezone.localdate()
    # Quick stats for overview cards
    month_start = today.replace(day=1)
    total_emp   = Employee.objects.filter(status='active').count()
    month_atts  = Attendance.objects.filter(date__gte=month_start, date__lte=today)
    return render(request, 'attendance_app/reports.html', {
        'today': today,
        'total_emp': total_emp,
        'month_present': month_atts.filter(status='present').count(),
        'month_absent':  month_atts.filter(status='absent').count(),
        'month_ot': float(month_atts.aggregate(s=__import__('django.db.models',fromlist=['Sum']).Sum('ot_hours'))['s'] or 0),
        'months': [(i, datetime.date(2000,i,1).strftime('%B')) for i in range(1,13)],
        'years':  range(2020, today.year+2),
    })


@login_required
def report_monthly(request):
    """Monthly employee-wise attendance summary."""
    if not has_perm(request.user, 'export_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    today = timezone.localdate()
    year, month, emp_id, date_from, date_to = _get_report_filters(request)
    employees = Employee.objects.filter(status='active').order_by('name')
    if emp_id:
        employees = employees.filter(pk=emp_id)

    holidays = {h.date for h in Holiday.objects.filter(date__gte=date_from, date__lte=date_to)}
    working_days = sum(
        1 for d in (date_from + datetime.timedelta(i) for i in range((date_to-date_from).days+1))
        if d not in holidays and d.weekday() != 6
    )

    att_map = {}
    for att in Attendance.objects.filter(date__gte=date_from, date__lte=date_to).select_related('employee'):
        att_map.setdefault(att.employee_id, []).append(att)

    rows = []
    for emp in employees:
        atts = att_map.get(emp.pk, [])
        present  = sum(1 for a in atts if a.status == 'present')
        half_day = sum(1 for a in atts if a.status == 'half_day')
        absent   = sum(1 for a in atts if a.status == 'absent')
        leave    = sum(1 for a in atts if a.status == 'leave')
        total_ot = sum(float(a.ot_hours) for a in atts)
        total_hrs = sum(float(a.total_hours()) for a in atts)
        unmarked  = max(0, working_days - present - half_day - absent - leave)
        att_rate  = round((present + half_day*0.5) / working_days * 100) if working_days else 0
        rows.append({
            'employee': emp,
            'present': present, 'half_day': half_day,
            'absent': absent + unmarked, 'leave': leave,
            'total_ot': round(total_ot, 1),
            'total_hrs': round(total_hrs, 1),
            'att_rate': att_rate,
            'working_days': working_days,
        })

    return render(request, 'attendance_app/report_monthly.html', {
        'rows': rows, 'year': year, 'month': month,
        'month_name': date_from.strftime('%B %Y'),
        'working_days': working_days,
        'employees': Employee.objects.filter(status='active').order_by('name'),
        'selected_emp': emp_id,
        'months': [(i, datetime.date(2000,i,1).strftime('%B')) for i in range(1,13)],
        'years':  range(2020, today.year+2),
        'date_from': date_from, 'date_to': date_to,
    })


@login_required
def report_monthly_export(request):
    """Export monthly summary to Excel."""
    if not has_perm(request.user, 'export_view'):
        return HttpResponse('Access denied', status=403)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    year, month, emp_id, date_from, date_to = _get_report_filters(request)
    employees = Employee.objects.filter(status='active').order_by('name')
    if emp_id:
        employees = employees.filter(pk=emp_id)

    holidays = {h.date for h in Holiday.objects.filter(date__gte=date_from, date__lte=date_to)}
    working_days = sum(
        1 for d in (date_from + datetime.timedelta(i) for i in range((date_to-date_from).days+1))
        if d not in holidays and d.weekday() != 6
    )
    att_map = {}
    for att in Attendance.objects.filter(date__gte=date_from, date__lte=date_to).select_related('employee'):
        att_map.setdefault(att.employee_id, []).append(att)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = f"Monthly {date_from.strftime('%b %Y')}"

    hdr_fill = PatternFill('solid', fgColor='E8650A')
    sub_fill = PatternFill('solid', fgColor='FFF3E8')
    bold_wh  = Font(bold=True, color='FFFFFF', size=11)
    bold_or  = Font(bold=True, color='E8650A', size=12)
    bold     = Font(bold=True, size=10)
    normal   = Font(size=10)
    center   = Alignment(horizontal='center', vertical='center')
    left     = Alignment(horizontal='left',   vertical='center')
    thin     = Border(*[Side(style='thin', color='CCCCCC')]*0,
                      left=Side(style='thin',color='CCCCCC'),
                      right=Side(style='thin',color='CCCCCC'),
                      top=Side(style='thin',color='CCCCCC'),
                      bottom=Side(style='thin',color='CCCCCC'))

    cs = CompanySettings.get_settings()
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = cs.company_name
    ws['A1'].font = bold_or
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:J2')
    ws['A2'] = f'Monthly Attendance Summary — {date_from.strftime("%B %Y")}   |   Working Days: {working_days}'
    ws['A2'].font = Font(size=10, color='666666')
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 16

    ws.append([])  # spacer

    headers = ['#','Employee','Job Title','Present','Half Day','Absent','Leave','Total Hours','OT Hours','Att. Rate %']
    ws.append(headers)
    hr = ws.max_row
    for ci, h in enumerate(headers, 1):
        c = ws.cell(hr, ci)
        c.value = h; c.font = bold_wh; c.fill = hdr_fill
        c.alignment = center; c.border = thin
    ws.row_dimensions[hr].height = 18

    for idx, emp in enumerate(employees, 1):
        atts = att_map.get(emp.pk, [])
        present  = sum(1 for a in atts if a.status=='present')
        half_day = sum(1 for a in atts if a.status=='half_day')
        absent   = sum(1 for a in atts if a.status=='absent')
        leave    = sum(1 for a in atts if a.status=='leave')
        total_ot = round(sum(float(a.ot_hours) for a in atts),1)
        total_hrs= round(sum(float(a.total_hours()) for a in atts),1)
        unmarked = max(0, working_days - present - half_day - absent - leave)
        att_rate = round((present + half_day*0.5)/working_days*100) if working_days else 0

        row = [idx, emp.name, emp.job_title or '', present, half_day,
               absent+unmarked, leave, total_hrs, total_ot, f'{att_rate}%']
        ws.append(row)
        rr = ws.max_row
        fill = PatternFill('solid', fgColor='FFFFFF') if idx%2==0 else PatternFill('solid', fgColor='FFF8F3')
        for ci, v in enumerate(row, 1):
            c = ws.cell(rr, ci)
            c.fill = fill; c.border = thin
            c.alignment = center if ci != 2 else left
            c.font = bold if ci in (1,10) else normal
            if ci == 10:
                rate = int(str(v).replace('%',''))
                c.font = Font(bold=True, size=10,
                    color='16A34A' if rate>=80 else 'D97706' if rate>=60 else 'DC2626')
        ws.row_dimensions[rr].height = 16

    col_widths = [5,28,20,10,10,10,10,13,10,12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    log_activity(request.user,'export','Report',description=f'Monthly report {date_from.strftime("%b %Y")}',request=request)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="monthly_report_{year}_{month:02d}.xlsx"'
    return resp


@login_required
def report_ot(request):
    """OT (overtime) report per employee."""
    if not has_perm(request.user, 'export_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    today = timezone.localdate()
    year, month, emp_id, date_from, date_to = _get_report_filters(request)
    employees = Employee.objects.filter(status='active').order_by('name')
    if emp_id:
        employees = employees.filter(pk=emp_id)

    att_qs = Attendance.objects.filter(
        date__gte=date_from, date__lte=date_to, ot_hours__gt=0
    ).select_related('employee').order_by('employee__name', 'date')

    # Group by employee
    from collections import defaultdict
    emp_ot = defaultdict(list)
    for a in att_qs:
        emp_ot[a.employee_id].append(a)

    rows = []
    for emp in employees:
        ot_records = emp_ot.get(emp.pk, [])
        total_ot   = round(sum(float(a.ot_hours) for a in ot_records), 1)
        ot_days    = len(ot_records)
        rows.append({
            'employee': emp,
            'total_ot': total_ot,
            'ot_days': ot_days,
            'records': ot_records,
            'avg_ot': round(total_ot/ot_days, 1) if ot_days else 0,
        })
    rows.sort(key=lambda x: -x['total_ot'])

    return render(request, 'attendance_app/report_ot.html', {
        'rows': rows, 'year': year, 'month': month,
        'month_name': date_from.strftime('%B %Y'),
        'employees': Employee.objects.filter(status='active').order_by('name'),
        'selected_emp': emp_id,
        'months': [(i, datetime.date(2000,i,1).strftime('%B')) for i in range(1,13)],
        'years':  range(2020, today.year+2),
        'grand_total_ot': round(sum(r['total_ot'] for r in rows), 1),
    })


@login_required
def report_ot_export(request):
    """Export OT report to Excel."""
    if not has_perm(request.user, 'export_view'):
        return HttpResponse('Access denied', status=403)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    year, month, emp_id, date_from, date_to = _get_report_filters(request)
    employees = Employee.objects.filter(status='active').order_by('name')
    if emp_id:
        employees = employees.filter(pk=emp_id)

    att_qs = Attendance.objects.filter(
        date__gte=date_from, date__lte=date_to, ot_hours__gt=0
    ).select_related('employee').order_by('date')

    from collections import defaultdict
    emp_ot = defaultdict(list)
    for a in att_qs:
        emp_ot[a.employee_id].append(a)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"OT {date_from.strftime('%b %Y')}"
    hfill = PatternFill('solid', fgColor='7C3AED')
    bold_wh = Font(bold=True, color='FFFFFF', size=10)
    bold = Font(bold=True, size=10); normal = Font(size=10)
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin = Border(left=Side(style='thin',color='CCCCCC'), right=Side(style='thin',color='CCCCCC'),
                  top=Side(style='thin',color='CCCCCC'), bottom=Side(style='thin',color='CCCCCC'))

    cs = CompanySettings.get_settings()
    ws.merge_cells('A1:F1')
    ws['A1'] = f'{cs.company_name} — OT Report — {date_from.strftime("%B %Y")}'
    ws['A1'].font = Font(bold=True, size=12, color='7C3AED'); ws['A1'].alignment = center
    ws.row_dimensions[1].height = 20; ws.append([])

    headers = ['#','Employee','Job Title','OT Days','Total OT Hours','Avg OT/Day']
    ws.append(headers)
    hr = ws.max_row
    for ci, h in enumerate(headers,1):
        c=ws.cell(hr,ci); c.value=h; c.font=bold_wh; c.fill=hfill; c.alignment=center; c.border=thin
    ws.row_dimensions[hr].height = 16

    for idx, emp in enumerate(employees, 1):
        recs = emp_ot.get(emp.pk, [])
        total_ot = round(sum(float(a.ot_hours) for a in recs), 1)
        ot_days  = len(recs)
        avg_ot   = round(total_ot/ot_days, 1) if ot_days else 0
        row = [idx, emp.name, emp.job_title or '', ot_days, total_ot, avg_ot]
        ws.append(row)
        rr = ws.max_row
        fill = PatternFill('solid', fgColor='FAF5FF' if idx%2 else 'FFFFFF')
        for ci, v in enumerate(row, 1):
            c=ws.cell(rr,ci); c.fill=fill; c.border=thin
            c.alignment=left if ci==2 else center; c.font=bold if ci==5 else normal
        ws.row_dimensions[rr].height = 15

    # Grand total
    ws.append([])
    gt = ws.max_row+1
    ws.cell(gt, 2, 'GRAND TOTAL').font = Font(bold=True, size=10, color='7C3AED')
    ws.cell(gt, 5, round(sum(sum(float(a.ot_hours) for a in emp_ot.get(e.pk,[])) for e in employees),1)).font = Font(bold=True, size=11, color='7C3AED')

    for ci, w in enumerate([5,28,20,10,15,12],1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    log_activity(request.user,'export','Report',description=f'OT report {date_from.strftime("%b %Y")}',request=request)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="ot_report_{year}_{month:02d}.xlsx"'
    return resp


@login_required
def report_absent(request):
    """Absent frequency report — who is absent most."""
    if not has_perm(request.user, 'export_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    today = timezone.localdate()
    year, month, emp_id, date_from, date_to = _get_report_filters(request)
    employees = Employee.objects.filter(status='active').order_by('name')
    if emp_id:
        employees = employees.filter(pk=emp_id)

    holidays = {h.date for h in Holiday.objects.filter(date__gte=date_from, date__lte=date_to)}
    working_days = sum(
        1 for d in (date_from + datetime.timedelta(i) for i in range((date_to-date_from).days+1))
        if d not in holidays and d.weekday() != 6
    )

    att_map = {}
    for att in Attendance.objects.filter(date__gte=date_from, date__lte=date_to).select_related('employee'):
        att_map.setdefault(att.employee_id, []).append(att)

    rows = []
    for emp in employees:
        atts   = att_map.get(emp.pk, [])
        absent = sum(1 for a in atts if a.status == 'absent')
        leave  = sum(1 for a in atts if a.status == 'leave')
        present= sum(1 for a in atts if a.status in ('present','half_day'))
        unmarked = max(0, working_days - len(atts))
        total_absent = absent + unmarked
        absent_dates = [a.date for a in atts if a.status=='absent']
        rows.append({
            'employee': emp,
            'absent': total_absent,
            'leave': leave,
            'present': present,
            'working_days': working_days,
            'absent_rate': round(total_absent/working_days*100) if working_days else 0,
            'absent_dates': sorted(absent_dates),
        })
    rows.sort(key=lambda x: -x['absent'])

    return render(request, 'attendance_app/report_absent.html', {
        'rows': rows, 'year': year, 'month': month,
        'month_name': date_from.strftime('%B %Y'),
        'working_days': working_days,
        'employees': Employee.objects.filter(status='active').order_by('name'),
        'selected_emp': emp_id,
        'months': [(i, datetime.date(2000,i,1).strftime('%B')) for i in range(1,13)],
        'years':  range(2020, today.year+2),
    })


@login_required
def report_absent_export(request):
    """Export absent frequency report to Excel."""
    if not has_perm(request.user, 'export_view'):
        return HttpResponse('Access denied', status=403)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    year, month, emp_id, date_from, date_to = _get_report_filters(request)
    employees = Employee.objects.filter(status='active').order_by('name')
    if emp_id:
        employees = employees.filter(pk=emp_id)
    holidays = {h.date for h in Holiday.objects.filter(date__gte=date_from, date__lte=date_to)}
    working_days = sum(
        1 for d in (date_from + datetime.timedelta(i) for i in range((date_to-date_from).days+1))
        if d not in holidays and d.weekday() != 6
    )
    att_map = {}
    for att in Attendance.objects.filter(date__gte=date_from, date__lte=date_to).select_related('employee'):
        att_map.setdefault(att.employee_id, []).append(att)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"Absent {date_from.strftime('%b %Y')}"
    hfill = PatternFill('solid', fgColor='DC2626')
    bold_wh = Font(bold=True, color='FFFFFF', size=10)
    bold = Font(bold=True, size=10); normal = Font(size=10)
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin = Border(left=Side(style='thin',color='CCCCCC'), right=Side(style='thin',color='CCCCCC'),
                  top=Side(style='thin',color='CCCCCC'), bottom=Side(style='thin',color='CCCCCC'))

    cs = CompanySettings.get_settings()
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{cs.company_name} — Absent Report — {date_from.strftime("%B %Y")}   (Working Days: {working_days})'
    ws['A1'].font = Font(bold=True, size=12, color='DC2626'); ws['A1'].alignment = center
    ws.row_dimensions[1].height = 20; ws.append([])

    headers = ['#','Employee','Job Title','Present Days','Absent Days','Leave Days','Absent Rate %']
    ws.append(headers)
    hr = ws.max_row
    for ci,h in enumerate(headers,1):
        c=ws.cell(hr,ci); c.value=h; c.font=bold_wh; c.fill=hfill; c.alignment=center; c.border=thin
    ws.row_dimensions[hr].height = 16

    for idx, emp in enumerate(employees,1):
        atts   = att_map.get(emp.pk, [])
        absent = sum(1 for a in atts if a.status=='absent')
        leave  = sum(1 for a in atts if a.status=='leave')
        present= sum(1 for a in atts if a.status in ('present','half_day'))
        unmarked = max(0, working_days - len(atts))
        total_absent = absent + unmarked
        rate = round(total_absent/working_days*100) if working_days else 0
        row = [idx, emp.name, emp.job_title or '', present, total_absent, leave, f'{rate}%']
        ws.append(row)
        rr = ws.max_row
        fill = PatternFill('solid', fgColor='FFF5F5' if idx%2 else 'FFFFFF')
        for ci, v in enumerate(row, 1):
            c=ws.cell(rr,ci); c.fill=fill; c.border=thin
            c.alignment=left if ci==2 else center; c.font=normal
            if ci==7:
                r=int(str(v).replace('%',''))
                c.font=Font(bold=True,size=10,color='DC2626' if r>=30 else 'D97706' if r>=15 else '16A34A')
        ws.row_dimensions[rr].height = 15

    for ci, w in enumerate([5,28,20,13,13,12,14],1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    log_activity(request.user,'export','Report',description=f'Absent report {date_from.strftime("%b %Y")}',request=request)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="absent_report_{year}_{month:02d}.xlsx"'
    return resp


@login_required
def report_late(request):
    """Late arrivals report — employees arriving after default in time."""
    if not has_perm(request.user, 'export_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    today = timezone.localdate()
    year, month, emp_id, date_from, date_to = _get_report_filters(request)

    cs = CompanySettings.get_settings()
    late_after_str = request.GET.get('late_after', cs.default_in_time.strftime('%H:%M'))
    try:
        late_after = datetime.time.fromisoformat(late_after_str)
    except ValueError:
        late_after = cs.default_in_time

    employees = Employee.objects.filter(status='active').order_by('name')
    if emp_id:
        employees = employees.filter(pk=emp_id)

    # All present/half-day records with in_time set
    att_qs = Attendance.objects.filter(
        date__gte=date_from, date__lte=date_to,
        status__in=['present','half_day'],
        in_time__isnull=False
    ).select_related('employee').order_by('employee__name','date')

    from collections import defaultdict
    emp_late = defaultdict(list)
    for a in att_qs:
        if a.in_time > late_after:
            delay_mins = (datetime.datetime.combine(a.date, a.in_time) -
                          datetime.datetime.combine(a.date, late_after)).seconds // 60
            emp_late[a.employee_id].append({
                'date': a.date,
                'in_time': a.in_time,
                'delay_mins': delay_mins,
            })

    rows = []
    for emp in employees:
        late_recs = emp_late.get(emp.pk, [])
        if late_recs or not emp_id:
            total_delay = sum(r['delay_mins'] for r in late_recs)
            rows.append({
                'employee': emp,
                'late_days': len(late_recs),
                'total_delay_mins': total_delay,
                'avg_delay_mins': round(total_delay/len(late_recs)) if late_recs else 0,
                'records': late_recs,
            })
    rows.sort(key=lambda x: -x['late_days'])

    return render(request, 'attendance_app/report_late.html', {
        'rows': rows, 'year': year, 'month': month,
        'month_name': date_from.strftime('%B %Y'),
        'late_after': late_after_str,
        'employees': Employee.objects.filter(status='active').order_by('name'),
        'selected_emp': emp_id,
        'months': [(i, datetime.date(2000,i,1).strftime('%B')) for i in range(1,13)],
        'years':  range(2020, today.year+2),
        'grand_total_late': sum(r['late_days'] for r in rows),
    })


@login_required
def report_late_export(request):
    """Export late arrivals report to Excel."""
    if not has_perm(request.user, 'export_view'):
        return HttpResponse('Access denied', status=403)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    year, month, emp_id, date_from, date_to = _get_report_filters(request)
    cs = CompanySettings.get_settings()
    late_after_str = request.GET.get('late_after', cs.default_in_time.strftime('%H:%M'))
    try:
        late_after = datetime.time.fromisoformat(late_after_str)
    except ValueError:
        late_after = cs.default_in_time

    employees = Employee.objects.filter(status='active').order_by('name')
    if emp_id:
        employees = employees.filter(pk=emp_id)

    att_qs = Attendance.objects.filter(
        date__gte=date_from, date__lte=date_to,
        status__in=['present','half_day'], in_time__isnull=False
    ).select_related('employee').order_by('employee__name','date')

    from collections import defaultdict
    emp_late = defaultdict(list)
    for a in att_qs:
        if a.in_time > late_after:
            delay_mins = (datetime.datetime.combine(a.date, a.in_time) -
                          datetime.datetime.combine(a.date, late_after)).seconds // 60
            emp_late[a.employee_id].append({'date': a.date, 'in_time': a.in_time, 'delay_mins': delay_mins})

    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"Late {date_from.strftime('%b %Y')}"
    hfill = PatternFill('solid', fgColor='D97706')
    bold_wh = Font(bold=True, color='FFFFFF', size=10)
    bold = Font(bold=True, size=10); normal = Font(size=10)
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    thin = Border(left=Side(style='thin',color='CCCCCC'), right=Side(style='thin',color='CCCCCC'),
                  top=Side(style='thin',color='CCCCCC'), bottom=Side(style='thin',color='CCCCCC'))

    ws.merge_cells('A1:F1')
    ws['A1'] = f'{cs.company_name} — Late Arrivals — {date_from.strftime("%B %Y")}   (After {late_after_str})'
    ws['A1'].font = Font(bold=True, size=12, color='D97706'); ws['A1'].alignment = center
    ws.row_dimensions[1].height = 20; ws.append([])

    headers = ['#','Employee','Job Title','Late Days','Total Delay (mins)','Avg Delay (mins)']
    ws.append(headers)
    hr = ws.max_row
    for ci,h in enumerate(headers,1):
        c=ws.cell(hr,ci); c.value=h; c.font=bold_wh; c.fill=hfill; c.alignment=center; c.border=thin
    ws.row_dimensions[hr].height = 16

    for idx, emp in enumerate(employees,1):
        recs = emp_late.get(emp.pk,[])
        total_delay = sum(r['delay_mins'] for r in recs)
        avg_delay   = round(total_delay/len(recs)) if recs else 0
        row = [idx, emp.name, emp.job_title or '', len(recs), total_delay, avg_delay]
        ws.append(row)
        rr = ws.max_row
        fill = PatternFill('solid', fgColor='FFFBEB' if idx%2 else 'FFFFFF')
        for ci, v in enumerate(row, 1):
            c=ws.cell(rr,ci); c.fill=fill; c.border=thin
            c.alignment=left if ci==2 else center; c.font=bold if ci==4 and len(recs)>5 else normal
        ws.row_dimensions[rr].height = 15

    for ci, w in enumerate([5,28,20,12,18,16],1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    log_activity(request.user,'export','Report',description=f'Late report {date_from.strftime("%b %Y")}',request=request)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="late_report_{year}_{month:02d}.xlsx"'
    return resp


# ─── EMPLOYEE SELF-SERVICE PORTAL ─────────────────────────────────────────────

def _get_portal_employee(request):
    """Return the Employee linked to the logged-in user, or None."""
    try:
        return request.user.employee_profile
    except Exception:
        return None


@login_required
def portal_dashboard(request):
    """Employee portal home — shows this month's summary + recent records."""
    emp = _get_portal_employee(request)
    if not emp:
        # Admin/staff with no employee link → redirect to admin dashboard
        return redirect('dashboard')

    today      = timezone.localdate()
    month_start = today.replace(day=1)
    month_end   = today.replace(day=calendar.monthrange(today.year, today.month)[1])

    holidays = {h.date for h in Holiday.objects.filter(date__gte=month_start, date__lte=month_end)}
    working_days = sum(
        1 for d in (month_start + datetime.timedelta(i)
                    for i in range((month_end - month_start).days + 1))
        if d not in holidays and d.weekday() != 6
    )

    month_atts = Attendance.objects.filter(employee=emp, date__gte=month_start, date__lte=month_end)
    present   = month_atts.filter(status='present').count()
    half_day  = month_atts.filter(status='half_day').count()
    absent    = month_atts.filter(status='absent').count()
    leave     = month_atts.filter(status='leave').count()
    total_ot  = float(sum(float(a.ot_hours) for a in month_atts))
    att_rate  = round((present + half_day * 0.5) / working_days * 100) if working_days else 0

    # Recent 10 records
    recent_atts = Attendance.objects.filter(employee=emp).order_by('-date')[:10]

    # Today's record
    today_att = Attendance.objects.filter(employee=emp, date=today).first()

    # Upcoming holidays (next 30 days)
    upcoming_holidays = Holiday.objects.filter(
        date__gte=today, date__lte=today + datetime.timedelta(days=30)
    ).order_by('date')[:5]

    # Last 30 days trend
    trend_start = today - datetime.timedelta(days=29)
    holiday_dates_30 = {h.date for h in Holiday.objects.filter(date__gte=trend_start, date__lte=today)}
    trend_atts_30 = {a.date: a.status for a in Attendance.objects.filter(
        employee=emp, date__gte=trend_start, date__lte=today
    )}
    trend_status = []
    trend_labels = []
    for i in range(30):
        d = trend_start + datetime.timedelta(days=i)
        if d in trend_atts_30:
            trend_status.append(trend_atts_30[d])
        elif d in holiday_dates_30:
            trend_status.append('holiday')
        elif d.weekday() == 6:
            trend_status.append('holiday')
        else:
            trend_status.append('unmarked')
        trend_labels.append(d.strftime('%d %b'))

    # total_hrs for the month
    total_hrs = round(float(sum(float(a.total_hours()) for a in month_atts)), 1)

    return render(request, 'attendance_app/portal_dashboard.html', {
        'emp': emp,
        'today': today,
        'today_att': today_att,
        'present': present, 'half_day': half_day,
        'absent': absent, 'leave': leave,
        'working_days': working_days,
        'att_rate': att_rate,
        'total_ot': round(total_ot, 1),
        'total_hrs': total_hrs,
        'recent_atts': recent_atts,
        'upcoming_holidays': upcoming_holidays,
        'month_name': today.strftime('%B %Y'),
        'trend_status': trend_status,
        'trend_labels': trend_labels,
    })


@login_required
def portal_attendance(request):
    """Full attendance history view with month/year filter + calendar."""
    emp = _get_portal_employee(request)
    if not emp:
        return redirect('dashboard')

    today = timezone.localdate()
    year  = int(request.GET.get('year',  today.year))
    month = int(request.GET.get('month', today.month))

    date_from = datetime.date(year, month, 1)
    date_to   = datetime.date(year, month, calendar.monthrange(year, month)[1])

    holidays     = Holiday.objects.filter(date__gte=date_from, date__lte=date_to)
    holiday_dates = {h.date: h.name for h in holidays}

    working_days = sum(
        1 for d in (date_from + datetime.timedelta(i)
                    for i in range((date_to - date_from).days + 1))
        if d not in holiday_dates and d.weekday() != 6
    )

    atts     = Attendance.objects.filter(employee=emp, date__gte=date_from, date__lte=date_to).order_by('date')
    att_dict = {a.date: a for a in atts}

    present  = sum(1 for a in atts if a.status == 'present')
    half_day = sum(1 for a in atts if a.status == 'half_day')
    absent   = sum(1 for a in atts if a.status == 'absent')
    leave    = sum(1 for a in atts if a.status == 'leave')
    total_ot = round(float(sum(float(a.ot_hours) for a in atts)), 1)
    total_hrs = round(float(sum(float(a.total_hours()) for a in atts)), 1)
    att_rate = round((present + half_day * 0.5) / working_days * 100) if working_days else 0

    # Build cal_days list for each day in the month (as expected by template)
    WEEKDAY_NAMES = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    cal_days = []
    num_days = calendar.monthrange(year, month)[1]
    for day_num in range(1, num_days + 1):
        d = datetime.date(year, month, day_num)
        is_holiday = d in holiday_dates
        is_future = d > today
        is_today_flag = d == today
        att = att_dict.get(d)
        cal_days.append({
            'date': d,
            'weekday': WEEKDAY_NAMES[d.weekday()],
            'is_holiday': is_holiday,
            'holiday_name': holiday_dates.get(d, ''),
            'is_future': is_future,
            'is_today': is_today_flag,
            'att': att,
        })

    return render(request, 'attendance_app/portal_attendance.html', {
        'emp': emp,
        'year': year, 'month': month,
        'month_name': date_from.strftime('%B %Y'),
        'date_from': date_from, 'date_to': date_to,
        'att_dict': att_dict,
        'holiday_dates': holiday_dates,
        'cal_days': cal_days,
        'today': today,
        'atts': atts,
        'present': present, 'half_day': half_day,
        'absent': absent, 'leave': leave,
        'total_ot': total_ot, 'total_hrs': total_hrs,
        'att_rate': att_rate, 'working_days': working_days,
        'months': [(i, datetime.date(2000, i, 1).strftime('%B')) for i in range(1, 13)],
        'years':  range(2020, today.year + 2),
    })


@login_required
def portal_month_data(request):
    """AJAX endpoint — returns attendance data for a given month as JSON."""
    emp = _get_portal_employee(request)
    if not emp:
        return JsonResponse({'error': 'No employee linked'}, status=403)

    year  = int(request.GET.get('year',  timezone.localdate().year))
    month = int(request.GET.get('month', timezone.localdate().month))
    date_from = datetime.date(year, month, 1)
    date_to   = datetime.date(year, month, calendar.monthrange(year, month)[1])

    atts = Attendance.objects.filter(employee=emp, date__gte=date_from, date__lte=date_to)
    data = {str(a.date): {
        'status': a.status,
        'in_time':  a.in_time.strftime('%H:%M')  if a.in_time  else '',
        'out_time': a.out_time.strftime('%H:%M') if a.out_time else '',
        'ot_hours': float(a.ot_hours),
        'total_hours': float(a.total_hours()),
    } for a in atts}
    return JsonResponse(data)


# ─── DOCUMENT MANAGEMENT ──────────────────────────────────────────────────────

@login_required
def document_list(request):
    if not has_perm(request.user, 'employee_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    today = timezone.localdate()
    emp_id   = request.GET.get('employee', '')
    type_id  = request.GET.get('doc_type', '')
    status_f = request.GET.get('status', '')
    q        = request.GET.get('q', '').strip()

    docs = EmployeeDocument.objects.select_related(
        'employee', 'doc_type', 'uploaded_by'
    ).filter(employee__status='active').order_by('expiry_date', 'employee__name')

    if emp_id:
        docs = docs.filter(employee_id=emp_id)
    if type_id:
        docs = docs.filter(doc_type_id=type_id)
    if q:
        docs = docs.filter(
            Q(employee__name__icontains=q) |
            Q(doc_number__icontains=q) |
            Q(doc_type__name__icontains=q)
        )

    # Annotate status in Python (property-based)
    all_docs = list(docs)
    if status_f:
        all_docs = [d for d in all_docs if d.status == status_f]

    # Counts for header pills
    valid_count    = sum(1 for d in all_docs if d.status == 'valid')
    expiring_count = sum(1 for d in all_docs if d.status == 'expiring')
    expired_count  = sum(1 for d in all_docs if d.status == 'expired')

    # Global expiry alerts (all employees, next 60 days)
    alert_docs = [d for d in EmployeeDocument.objects.select_related('employee','doc_type')
                  .filter(employee__status='active', expiry_date__isnull=False)
                  if d.status in ('expiring', 'expired')]

    return render(request, 'attendance_app/document_list.html', {
        'docs': all_docs,
        'employees': Employee.objects.filter(status='active').order_by('name'),
        'doc_types': DocumentType.objects.all(),
        'selected_emp': emp_id,
        'selected_type': type_id,
        'selected_status': status_f,
        'q': q,
        'valid_count': valid_count,
        'expiring_count': expiring_count,
        'expired_count': expired_count,
        'alert_count': len(alert_docs),
        'today': today,
    })


@login_required
def document_expiring(request):
    """Focused view: expiring + expired documents across all employees."""
    if not has_perm(request.user, 'employee_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    today = timezone.localdate()
    all_docs = list(EmployeeDocument.objects.select_related('employee', 'doc_type')
                    .filter(employee__status='active', expiry_date__isnull=False)
                    .order_by('expiry_date'))

    expired  = [d for d in all_docs if d.status == 'expired']
    expiring = [d for d in all_docs if d.status == 'expiring']

    return render(request, 'attendance_app/document_expiring.html', {
        'expired': expired,
        'expiring': expiring,
        'today': today,
    })


@login_required
def document_add(request):
    if not has_perm(request.user, 'employee_edit'):
        messages.error(request, 'Access denied.')
        return redirect('document_list')

    employees  = Employee.objects.filter(status='active').order_by('name')
    doc_types  = DocumentType.objects.all()
    preselect_emp = request.GET.get('employee', '')

    if not doc_types.exists():
        messages.warning(request, 'Please create at least one Document Type first.')
        return redirect('doctype_list')

    if request.method == 'POST':
        emp_id  = request.POST.get('employee')
        type_id = request.POST.get('doc_type')
        try:
            emp      = Employee.objects.get(pk=emp_id)
            doc_type = DocumentType.objects.get(pk=type_id)
        except (Employee.DoesNotExist, DocumentType.DoesNotExist):
            messages.error(request, 'Invalid employee or document type.')
            return redirect('document_add')

        doc = EmployeeDocument(
            employee    = emp,
            doc_type    = doc_type,
            doc_number  = request.POST.get('doc_number', '').strip() or None,
            notes       = request.POST.get('notes', '').strip() or None,
            uploaded_by = request.user,
        )
        issue_str  = request.POST.get('issue_date', '').strip()
        expiry_str = request.POST.get('expiry_date', '').strip()
        if issue_str:
            doc.issue_date = datetime.date.fromisoformat(issue_str)
        if expiry_str:
            doc.expiry_date = datetime.date.fromisoformat(expiry_str)
        if request.FILES.get('file'):
            doc.file = request.FILES['file']
        doc.save()

        log_activity(request.user, 'create', 'Document', doc,
                     f'Added {doc_type.name} for {emp.name}', request)
        messages.success(request, f'{doc_type.name} added for {emp.name}.')
        return redirect(f'/documents/?employee={emp_id}')

    return render(request, 'attendance_app/document_form.html', {
        'action': 'Add',
        'employees': employees,
        'doc_types': doc_types,
        'preselect_emp': preselect_emp,
    })


@login_required
def document_edit(request, pk):
    if not has_perm(request.user, 'employee_edit'):
        messages.error(request, 'Access denied.')
        return redirect('document_list')

    doc       = get_object_or_404(EmployeeDocument, pk=pk)
    employees = Employee.objects.filter(status='active').order_by('name')
    doc_types = DocumentType.objects.all()

    if request.method == 'POST':
        type_id = request.POST.get('doc_type')
        try:
            doc.doc_type = DocumentType.objects.get(pk=type_id)
        except DocumentType.DoesNotExist:
            pass

        doc.doc_number = request.POST.get('doc_number', '').strip() or None
        doc.notes      = request.POST.get('notes', '').strip() or None

        issue_str  = request.POST.get('issue_date', '').strip()
        expiry_str = request.POST.get('expiry_date', '').strip()
        doc.issue_date  = datetime.date.fromisoformat(issue_str)  if issue_str  else None
        doc.expiry_date = datetime.date.fromisoformat(expiry_str) if expiry_str else None

        if request.FILES.get('file'):
            doc.file = request.FILES['file']
        elif request.POST.get('remove_file') == '1':
            doc.file = None

        doc.save()
        log_activity(request.user, 'update', 'Document', doc,
                     f'Updated {doc.doc_type.name} for {doc.employee.name}', request)
        messages.success(request, 'Document updated.')
        return redirect(f'/documents/?employee={doc.employee_id}')

    return render(request, 'attendance_app/document_form.html', {
        'action': 'Edit',
        'doc': doc,
        'employees': employees,
        'doc_types': doc_types,
    })


@login_required
def document_delete(request, pk):
    if not has_perm(request.user, 'employee_delete'):
        messages.error(request, 'Access denied.')
        return redirect('document_list')
    doc = get_object_or_404(EmployeeDocument, pk=pk)
    emp_id = doc.employee_id
    if request.method == 'POST':
        desc = f'Deleted {doc.doc_type.name} for {doc.employee.name}'
        doc.delete()
        log_activity(request.user, 'delete', 'Document', description=desc, request=request)
        messages.success(request, 'Document deleted.')
    return redirect(f'/documents/?employee={emp_id}')


@login_required
def document_download(request, pk):
    """Serve document file for download."""
    if not has_perm(request.user, 'employee_view'):
        return HttpResponse('Access denied', status=403)
    doc = get_object_or_404(EmployeeDocument, pk=pk)
    if not doc.file:
        messages.error(request, 'No file attached to this document.')
        return redirect('document_list')
    import mimetypes, os
    file_path = doc.file.path
    mime, _ = mimetypes.guess_type(file_path)
    mime = mime or 'application/octet-stream'
    with open(file_path, 'rb') as f:
        resp = HttpResponse(f.read(), content_type=mime)
    filename = os.path.basename(file_path)
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# ── Document Types ─────────────────────────────────────────────────────────────

@login_required
def doctype_list(request):
    if not has_perm(request.user, 'employee_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    types = DocumentType.objects.annotate(doc_count=Count('employeedocument')).order_by('name')
    return render(request, 'attendance_app/doctype_list.html', {'types': types})


@login_required
def doctype_add(request):
    if not request.user.is_superuser:
        messages.error(request, 'Only administrators can manage document types.')
        return redirect('doctype_list')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Name is required.')
            return redirect('doctype_add')
        DocumentType.objects.create(
            name=name,
            description=request.POST.get('description', '').strip() or None,
            requires_expiry=request.POST.get('requires_expiry') == '1',
            alert_days=int(request.POST.get('alert_days', 30) or 30),
        )
        messages.success(request, f'Document type "{name}" created.')
        return redirect('doctype_list')
    return render(request, 'attendance_app/doctype_form.html', {'action': 'Add'})


@login_required
def doctype_edit(request, pk):
    if not request.user.is_superuser:
        messages.error(request, 'Only administrators can manage document types.')
        return redirect('doctype_list')
    dt = get_object_or_404(DocumentType, pk=pk)
    if request.method == 'POST':
        dt.name = request.POST.get('name', dt.name).strip()
        dt.description = request.POST.get('description', '').strip() or None
        dt.requires_expiry = request.POST.get('requires_expiry') == '1'
        dt.alert_days = int(request.POST.get('alert_days', 30) or 30)
        dt.save()
        messages.success(request, f'Document type "{dt.name}" updated.')
        return redirect('doctype_list')
    return render(request, 'attendance_app/doctype_form.html', {'action': 'Edit', 'dt': dt})


@login_required
def doctype_delete(request, pk):
    if not request.user.is_superuser:
        messages.error(request, 'Only administrators can delete document types.')
        return redirect('doctype_list')
    dt = get_object_or_404(DocumentType, pk=pk)
    if request.method == 'POST':
        try:
            name = dt.name
            dt.delete()
            messages.success(request, f'Document type "{name}" deleted.')
        except Exception:
            messages.error(request, 'Cannot delete — documents exist with this type.')
    return redirect('doctype_list')


# ─── BULK IMPORT ──────────────────────────────────────────────────────────────

@login_required
def import_hub(request):
    """Landing page — choose what to import."""
    if not has_perm(request.user, 'employee_add'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    # Quick stats
    total_emp = Employee.objects.count()
    total_att = Attendance.objects.count()

    return render(request, 'attendance_app/import_hub.html', {
        'total_emp': total_emp,
        'total_att': total_att,
    })


# ── Employee Import ────────────────────────────────────────────────────────────

@login_required
def import_employees_template(request):
    """Download a blank employee import template .xlsx"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employees'

    hfill  = PatternFill('solid', fgColor='E8650A')
    ifill  = PatternFill('solid', fgColor='FFF3E8')
    hfont  = Font(bold=True, color='FFFFFF', size=10)
    ifont  = Font(italic=True, color='999999', size=9)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),  bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = [
        'name *', 'emirates_id', 'mobile', 'dob\n(YYYY-MM-DD)',
        'joining_date\n(YYYY-MM-DD)', 'job_title', 'country',
        'address', 'emp_type\n(permanent/temporary)', 'status\n(active/disabled)',
    ]
    instructions = [
        'Full name — REQUIRED', 'Emirates ID (optional)', 'Mobile number',
        'Date of birth e.g. 1990-05-15', 'Joining date e.g. 2023-01-01',
        'Job title e.g. Engineer', 'Country e.g. India',
        'Full address', 'permanent OR temporary', 'active OR disabled',
    ]
    example = [
        'Ahmed Al Rashidi', '784-1990-1234567-1', '+971501234567',
        '1990-05-15', '2023-01-01', 'Software Engineer', 'UAE',
        'Dubai, UAE', 'permanent', 'active',
    ]

    col_widths = [22, 22, 16, 14, 14, 20, 14, 24, 16, 12]

    # Title row
    ws.merge_cells('A1:J1')
    ws['A1'] = 'EMPLOYEE IMPORT TEMPLATE — AttendPro'
    ws['A1'].font = Font(bold=True, size=12, color='E8650A')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # Header row
    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.font = hfont; c.fill = hfill; c.alignment = center; c.border = thin
    ws.row_dimensions[2].height = 30

    # Instruction row
    for ci, txt in enumerate(instructions, 1):
        c = ws.cell(3, ci, txt)
        c.font = ifont; c.fill = ifill; c.alignment = left; c.border = thin
    ws.row_dimensions[3].height = 14

    # Example row
    for ci, val in enumerate(example, 1):
        c = ws.cell(4, ci, val)
        c.font = Font(color='555555', size=10); c.alignment = left; c.border = thin
    ws.row_dimensions[4].height = 14

    # Empty data rows
    for r in range(5, 105):
        for ci in range(1, 11):
            ws.cell(r, ci).border = thin
        ws.row_dimensions[r].height = 14

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="employee_import_template.xlsx"'
    return resp


@login_required
def import_employees(request):
    """Upload, validate, preview, and confirm employee import."""
    if not has_perm(request.user, 'employee_add'):
        messages.error(request, 'Access denied.')
        return redirect('import_hub')

    if request.method == 'GET':
        return render(request, 'attendance_app/import_employees.html')

    # ── STEP 1: upload + validate ──────────────────────────────────────────────
    if 'file' in request.FILES and 'confirm' not in request.POST:
        f = request.FILES['file']
        if not f.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload an Excel file (.xlsx or .xls).')
            return render(request, 'attendance_app/import_employees.html')

        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
        except Exception as e:
            messages.error(request, f'Could not read Excel file: {e}')
            return render(request, 'attendance_app/import_employees.html')

        VALID_STATUS   = {'active', 'disabled'}
        VALID_EMP_TYPE = {'permanent', 'temporary'}

        rows_ok    = []
        rows_error = []
        rows_dup   = []

        # Skip rows 1-3 (title, header, instructions), data from row 4
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            if not any(row):
                continue  # blank row

            def cell(i):
                v = row[i] if i < len(row) else None
                return str(v).strip() if v is not None else ''

            name        = cell(0)
            emirates_id = cell(1)
            mobile      = cell(2)
            dob_s       = cell(3)
            join_s      = cell(4)
            job_title   = cell(5)
            country     = cell(6)
            address     = cell(7)
            emp_type    = cell(8).lower() or 'permanent'
            status      = cell(9).lower() or 'active'

            errs = []
            if not name:
                errs.append('Name is required')
            if emp_type and emp_type not in VALID_EMP_TYPE:
                errs.append(f'emp_type must be permanent or temporary (got "{emp_type}")')
            if status and status not in VALID_STATUS:
                errs.append(f'status must be active or disabled (got "{status}")')

            def parse_excel_date(raw, field_name, errs):
                if raw is None or raw == '': return None
                if hasattr(raw, 'date'): return raw.date()
                if hasattr(raw, 'year'): return raw
                s = str(raw).strip().split()[0]
                try: return datetime.date.fromisoformat(s)
                except ValueError:
                    errs.append(f'{field_name} format invalid (use YYYY-MM-DD, got "{s}")')
                    return None

            raw_dob  = row[3] if len(row) > 3 else None
            raw_join = row[4] if len(row) > 4 else None
            dob          = parse_excel_date(raw_dob,  'dob',          errs)
            joining_date = parse_excel_date(raw_join, 'joining_date', errs)

            record = {
                'row': row_idx, 'name': name, 'emirates_id': emirates_id or None,
                'mobile': mobile or None, 'dob': dob, 'joining_date': joining_date,
                'job_title': job_title or None, 'country': country or None,
                'address': address or None, 'emp_type': emp_type or 'permanent',
                'status': status or 'active',
            }

            if errs:
                record['errors'] = errs
                rows_error.append(record)
            else:
                # Duplicate check by name
                exists = Employee.objects.filter(name__iexact=name).first()
                if exists:
                    record['existing_id'] = exists.pk
                    rows_dup.append(record)
                else:
                    rows_ok.append(record)

        if not rows_ok and not rows_dup and not rows_error:
            messages.warning(request, 'No data rows found in the file.')
            return render(request, 'attendance_app/import_employees.html')

        # Store validated data in session for confirm step
        import json

        def serial(r):
            d = dict(r)
            if d.get('dob'):         d['dob']          = str(d['dob'])
            if d.get('joining_date'): d['joining_date'] = str(d['joining_date'])
            return d

        request.session['import_emp_ok']    = [serial(r) for r in rows_ok]
        request.session['import_emp_dup']   = [serial(r) for r in rows_dup]
        request.session['import_emp_error'] = [serial(r) for r in rows_error]

        return render(request, 'attendance_app/import_employees.html', {
            'preview': True,
            'rows_ok':    rows_ok,
            'rows_dup':   rows_dup,
            'rows_error': rows_error,
        })

    # ── STEP 2: confirm + save ─────────────────────────────────────────────────
    if 'confirm' in request.POST:
        rows_ok  = request.session.pop('import_emp_ok',  [])
        rows_dup = request.session.pop('import_emp_dup', [])
        request.session.pop('import_emp_error', None)

        overwrite = request.POST.get('overwrite_duplicates') == '1'
        created = updated = skipped = 0

        def parse_date(s):
            try: return datetime.date.fromisoformat(s) if s else None
            except: return None

        for r in rows_ok:
            Employee.objects.create(
                name=r['name'], emirates_id=r['emirates_id'], mobile=r['mobile'],
                dob=parse_date(r['dob']), joining_date=parse_date(r['joining_date']),
                job_title=r['job_title'], country=r['country'], address=r['address'],
                emp_type=r['emp_type'], status=r['status'],
            )
            created += 1

        for r in rows_dup:
            if overwrite:
                Employee.objects.filter(pk=r['existing_id']).update(
                    emirates_id=r['emirates_id'], mobile=r['mobile'],
                    dob=parse_date(r['dob']), joining_date=parse_date(r['joining_date']),
                    job_title=r['job_title'], country=r['country'], address=r['address'],
                    emp_type=r['emp_type'], status=r['status'],
                )
                updated += 1
            else:
                skipped += 1

        log_activity(request.user, 'create', 'Employee',
                     description=f'Bulk import: {created} created, {updated} updated, {skipped} skipped',
                     request=request)

        parts = []
        if created:  parts.append(f'{created} employee{"s" if created!=1 else ""} created')
        if updated:  parts.append(f'{updated} updated')
        if skipped:  parts.append(f'{skipped} skipped (duplicate)')
        messages.success(request, 'Import complete — ' + ', '.join(parts) + '.')
        return redirect('employee_list')

    return render(request, 'attendance_app/import_employees.html')


# ── Attendance Import ──────────────────────────────────────────────────────────

@login_required
def import_attendance_template(request):
    """Download a blank attendance import template .xlsx"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    hfill  = PatternFill('solid', fgColor='2563EB')
    ifill  = PatternFill('solid', fgColor='EFF6FF')
    hfont  = Font(bold=True, color='FFFFFF', size=10)
    ifont  = Font(italic=True, color='999999', size=9)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center')
    thin   = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),  bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = [
        'employee_name *', 'date *\n(YYYY-MM-DD)',
        'status *\n(present/absent/half_day/holiday/leave)',
        'in_time\n(HH:MM)', 'out_time\n(HH:MM)',
        'ot_hours\n(e.g. 1.5)', 'notes',
    ]
    instructions = [
        'Must match employee name exactly', 'e.g. 2025-03-01 — REQUIRED',
        'present / absent / half_day / holiday / leave — REQUIRED',
        'e.g. 07:00 (optional)', 'e.g. 17:00 (optional)',
        'Overtime hours e.g. 1.5 (optional)', 'Any notes (optional)',
    ]
    example_rows = [
        ['Ahmed Al Rashidi', '2025-03-01', 'present', '07:05', '17:10', '1.5', ''],
        ['Fatima Hassan',    '2025-03-01', 'absent',  '',      '',       '',    'Sick leave'],
        ['John Smith',       '2025-03-01', 'half_day','07:00', '12:00',  '',    ''],
    ]

    col_widths = [26, 14, 32, 10, 10, 10, 24]

    ws.merge_cells('A1:G1')
    ws['A1'] = 'ATTENDANCE IMPORT TEMPLATE — AttendPro'
    ws['A1'].font = Font(bold=True, size=12, color='2563EB')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.font = hfont; c.fill = hfill; c.alignment = center; c.border = thin
    ws.row_dimensions[2].height = 30

    for ci, txt in enumerate(instructions, 1):
        c = ws.cell(3, ci, txt)
        c.font = ifont; c.fill = ifill; c.alignment = left; c.border = thin
    ws.row_dimensions[3].height = 14

    for ei, ex in enumerate(example_rows, 4):
        for ci, val in enumerate(ex, 1):
            c = ws.cell(ei, ci, val)
            c.font = Font(color='555555', size=10); c.alignment = left; c.border = thin
        ws.row_dimensions[ei].height = 14

    for r in range(7, 507):
        for ci in range(1, 8):
            ws.cell(r, ci).border = thin
        ws.row_dimensions[r].height = 14

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="attendance_import_template.xlsx"'
    return resp


@login_required
def import_attendance(request):
    """Upload, validate, preview, and confirm attendance import."""
    if not has_perm(request.user, 'attendance_add'):
        messages.error(request, 'Access denied.')
        return redirect('import_hub')

    if request.method == 'GET':
        return render(request, 'attendance_app/import_attendance.html')

    # ── STEP 1: upload + validate ──────────────────────────────────────────────
    if 'file' in request.FILES and 'confirm' not in request.POST:
        f = request.FILES['file']
        if not f.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload an Excel file (.xlsx or .xls).')
            return render(request, 'attendance_app/import_attendance.html')

        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
        except Exception as e:
            messages.error(request, f'Could not read Excel file: {e}')
            return render(request, 'attendance_app/import_attendance.html')

        VALID_STATUS = {'present', 'absent', 'half_day', 'holiday', 'leave'}

        # Build employee lookup (name → Employee) — case-insensitive
        emp_map = {e.name.lower(): e for e in Employee.objects.filter(status='active')}

        rows_ok    = []
        rows_error = []
        rows_dup   = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            if not any(row):
                continue

            def cell(i, default=''):
                v = row[i] if i < len(row) else None
                return str(v).strip() if v is not None else default

            emp_name = cell(0)
            date_s   = cell(1)
            status   = cell(2).lower()
            in_s     = cell(3)
            out_s    = cell(4)
            ot_s     = cell(5)
            notes    = cell(6)

            errs = []
            emp_obj = None

            if not emp_name:
                errs.append('employee_name is required')
            else:
                emp_obj = emp_map.get(emp_name.lower())
                if not emp_obj:
                    errs.append(f'Employee "{emp_name}" not found (check spelling / active status)')

            att_date = None
            # openpyxl may return a datetime/date object directly instead of string
            raw_date = row[1] if len(row) > 1 else None
            if raw_date is None or raw_date == '':
                errs.append('date is required')
            elif hasattr(raw_date, 'date'):
                # It's a datetime object — extract date part directly
                att_date = raw_date.date()
            elif hasattr(raw_date, 'year'):
                # It's already a date object
                att_date = raw_date
            else:
                try:
                    att_date = datetime.date.fromisoformat(str(raw_date).strip().split()[0])
                except ValueError:
                    errs.append(f'date format invalid — use YYYY-MM-DD (got "{date_s}")')

            if not status:
                errs.append('status is required')
            elif status not in VALID_STATUS:
                errs.append(f'status must be present/absent/half_day/holiday/leave (got "{status}")')

            def parse_excel_time(raw, field_name, errs):
                if raw is None or raw == '': return None
                if isinstance(raw, datetime.time): return raw
                if isinstance(raw, datetime.datetime): return raw.time()
                s = str(raw).strip()
                try: return datetime.time.fromisoformat(s[:5])  # handles HH:MM or HH:MM:SS
                except ValueError:
                    errs.append(f'{field_name} format invalid — use HH:MM (got "{s}")')
                    return None

            raw_in  = row[3] if len(row) > 3 else None
            raw_out = row[4] if len(row) > 4 else None
            in_time  = parse_excel_time(raw_in,  'in_time',  errs)
            out_time = parse_excel_time(raw_out, 'out_time', errs)

            ot_hours = 0
            if ot_s:
                try:
                    ot_hours = float(ot_s)
                    if ot_hours < 0 or ot_hours > 24:
                        raise ValueError
                except ValueError:
                    errs.append(f'ot_hours must be a number 0-24 (got "{ot_s}")')

            record = {
                'row': row_idx,
                'emp_name': emp_name,
                'emp_id': emp_obj.pk if emp_obj else None,
                'date': str(att_date) if att_date else date_s,
                'status': status,
                'in_time': str(in_time) if in_time else '',
                'out_time': str(out_time) if out_time else '',
                'ot_hours': ot_hours,
                'notes': notes,
            }

            if errs:
                record['errors'] = errs
                rows_error.append(record)
            elif emp_obj and att_date:
                existing = Attendance.objects.filter(employee=emp_obj, date=att_date).first()
                if existing:
                    record['existing_id'] = existing.pk
                    record['existing_status'] = existing.status
                    rows_dup.append(record)
                else:
                    rows_ok.append(record)

        if not rows_ok and not rows_dup and not rows_error:
            messages.warning(request, 'No data rows found in the file.')
            return render(request, 'attendance_app/import_attendance.html')

        request.session['import_att_ok']    = rows_ok
        request.session['import_att_dup']   = rows_dup
        request.session['import_att_error'] = rows_error

        return render(request, 'attendance_app/import_attendance.html', {
            'preview': True,
            'rows_ok':    rows_ok,
            'rows_dup':   rows_dup,
            'rows_error': rows_error,
        })

    # ── STEP 2: confirm + save ─────────────────────────────────────────────────
    if 'confirm' in request.POST:
        rows_ok  = request.session.pop('import_att_ok',  [])
        rows_dup = request.session.pop('import_att_dup', [])
        request.session.pop('import_att_error', None)

        overwrite = request.POST.get('overwrite_duplicates') == '1'
        created = updated = skipped = 0

        def pt(s):
            try: return datetime.time.fromisoformat(s) if s else None
            except: return None

        for r in rows_ok:
            Attendance.objects.create(
                employee_id=r['emp_id'],
                date=datetime.date.fromisoformat(r['date']),
                status=r['status'],
                in_time=pt(r['in_time']),
                out_time=pt(r['out_time']),
                ot_hours=r['ot_hours'],
                notes=r['notes'] or None,
                created_by=request.user,
            )
            created += 1

        for r in rows_dup:
            if overwrite:
                Attendance.objects.filter(pk=r['existing_id']).update(
                    status=r['status'],
                    in_time=pt(r['in_time']),
                    out_time=pt(r['out_time']),
                    ot_hours=r['ot_hours'],
                    notes=r['notes'] or None,
                )
                updated += 1
            else:
                skipped += 1

        log_activity(request.user, 'create', 'Attendance',
                     description=f'Bulk import: {created} created, {updated} updated, {skipped} skipped',
                     request=request)

        parts = []
        if created:  parts.append(f'{created} record{"s" if created!=1 else ""} imported')
        if updated:  parts.append(f'{updated} updated')
        if skipped:  parts.append(f'{skipped} skipped (duplicate)')
        messages.success(request, 'Import complete — ' + ', '.join(parts) + '.')
        return redirect('attendance_list')

    return render(request, 'attendance_app/import_attendance.html')
