import os

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Q, Count
from django.conf import settings
from .models import (Employee, Attendance, Holiday, LeaveType, LeaveRequest,
                     UserPermission, ActivityLog, CompanySettings)
from .middleware import log_activity
from .decorators import permission_required
import datetime
import json
import calendar
import io

# ─── AUTH ────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            log_activity(user, 'login', 'Auth', description=f'{user.username} logged in', request=request)
            return redirect('dashboard')
        messages.error(request, 'Invalid username or password.')
    return render(request, 'attendance_app/login.html')


def logout_view(request):
    if request.user.is_authenticated:
        log_activity(request.user, 'logout', 'Auth', description=f'{request.user.username} logged out', request=request)
    logout(request)
    return redirect('login')


# ─── PERMISSION HELPER ────────────────────────────────────────────────────────

def has_perm(user, perm):
    if user.is_superuser:
        return True
    try:
        up = UserPermission.objects.get(user=user)
        return perm in up.permissions
    except UserPermission.DoesNotExist:
        return False


# ─── DASHBOARD ───────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    if not has_perm(request.user, 'dashboard_view') and not request.user.is_superuser:
        return redirect_to_first_permitted(request)

    today = timezone.localdate()
    total_employees = Employee.objects.filter(status="active").count()
    today_atts = list(Attendance.objects.filter(date=today).select_related("employee"))
    today_attendance = sum(1 for a in today_atts if a.status == "present")
    today_absent     = sum(1 for a in today_atts if a.status == "absent")
    today_half       = sum(1 for a in today_atts if a.status == "half_day")
    today_leave      = sum(1 for a in today_atts if a.status == "leave")
    today_total_marked = len(today_atts)
    today_not_marked = max(0, total_employees - today_total_marked)

    upcoming_holidays = Holiday.objects.filter(date__gte=today).order_by("date")[:5]
    recent_activity   = ActivityLog.objects.select_related("user").all()[:8]
    pending_leaves    = LeaveRequest.objects.filter(status="pending").count()
    approved_leaves   = LeaveRequest.objects.filter(
        status="approved", start_date__lte=today, end_date__gte=today
    ).count()

    # ── Month summary ──────────────────────────────────────────────────────
    month_start = today.replace(day=1)
    month_atts = Attendance.objects.filter(date__gte=month_start, date__lte=today)
    month_present  = month_atts.filter(status="present").count()
    month_absent   = month_atts.filter(status="absent").count()
    month_leave    = month_atts.filter(status="leave").count()
    month_half     = month_atts.filter(status="half_day").count()

    # ── 30-day trend data for chart ────────────────────────────────────────
    trend_labels, trend_present, trend_absent = [], [], []
    for i in range(29, -1, -1):
        d = today - datetime.timedelta(days=i)
        day_atts = Attendance.objects.filter(date=d)
        trend_labels.append(d.strftime("%d %b"))
        trend_present.append(day_atts.filter(status="present").count())
        trend_absent.append(day_atts.filter(status="absent").count())

    # ── Top employees this month (most present days) ───────────────────────
    from django.db.models import Count as DCount
    top_employees = (
        Attendance.objects.filter(date__gte=month_start, date__lte=today, status="present")
        .values("employee__name")
        .annotate(present_days=DCount("id"))
        .order_by("-present_days")[:5]
    )

    # ── Today absent employees ─────────────────────────────────────────────
    today_absent_emps = [a.employee for a in today_atts if a.status == "absent"][:6]

    # ── Next 7 days events ─────────────────────────────────────────────────
    week_end = today + datetime.timedelta(days=7)
    week_holidays = Holiday.objects.filter(date__gte=today, date__lte=week_end).order_by("date")

    context = {
        "total_employees": total_employees,
        "today_attendance": today_attendance,
        "today_absent": today_absent,
        "today_half": today_half,
        "today_leave": today_leave,
        "today_total_marked": today_total_marked,
        "today_not_marked": today_not_marked,
        "upcoming_holidays": upcoming_holidays,
        "week_holidays": week_holidays,
        "recent_activity": recent_activity,
        "pending_leaves": pending_leaves,
        "approved_leaves": approved_leaves,
        "today": today,
        "month_start": month_start,
        "month_present": month_present,
        "month_absent": month_absent,
        "month_leave": month_leave,
        "month_half": month_half,
        "trend_labels": trend_labels,
        "trend_present": trend_present,
        "trend_absent": trend_absent,
        "top_employees": list(top_employees),
        "today_absent_emps": today_absent_emps,
    }
    return render(request, "attendance_app/dashboard.html", context)


def redirect_to_first_permitted(request):
    perms_pages = [
        ('employee_view', 'employee_list'),
        ('attendance_view', 'attendance_list'),
        ('holiday_view', 'holiday_list'),
        ('leave_view', 'leave_list'),
        ('export_view', 'export_attendance'),
        ('activity_view', 'activity_log'),
    ]
    for perm, url_name in perms_pages:
        if has_perm(request.user, perm):
            return redirect(url_name)
    return render(request, 'attendance_app/no_access.html')


# ─── EMPLOYEES ───────────────────────────────────────────────────────────────

@login_required
def employee_list(request):
    if not has_perm(request.user, 'employee_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    employees = Employee.objects.all()
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    type_filter = request.GET.get('type', '')

    if search:
        employees = employees.filter(
            Q(name__icontains=search) | Q(job_title__icontains=search) |
            Q(emirates_id__icontains=search) | Q(mobile__icontains=search)
        )
    if status_filter:
        employees = employees.filter(status=status_filter)
    if type_filter:
        employees = employees.filter(emp_type=type_filter)

    return render(request, 'attendance_app/employee_list.html', {
        'employees': employees,
        'search': search,
        'status_filter': status_filter,
        'type_filter': type_filter,
    })


@login_required
def employee_add(request):
    if not has_perm(request.user, 'employee_add'):
        messages.error(request, 'Access denied.')
        return redirect('employee_list')

    if request.method == 'POST':
        emp = Employee()
        emp.name = request.POST.get('name', '').strip()
        if not emp.name:
            messages.error(request, 'Employee name is required.')
            return render(request, 'attendance_app/employee_form.html', {'form_data': request.POST})

        emp.emirates_id = request.POST.get('emirates_id') or None
        emp.mobile = request.POST.get('mobile') or None
        emp.dob = request.POST.get('dob') or None
        emp.joining_date = request.POST.get('joining_date') or None
        emp.job_title = request.POST.get('job_title') or None
        emp.country = request.POST.get('country') or None
        emp.address = request.POST.get('address') or None
        emp.description = request.POST.get('description') or None
        emp.emp_type = request.POST.get('emp_type', 'permanent')
        emp.status = request.POST.get('status', 'active')
        if request.FILES.get('photo'):
            emp.photo = request.FILES['photo']
        emp.save()

        log_activity(request.user, 'create', 'Employee', emp, f'Added employee {emp.name}', request)
        messages.success(request, f'Employee {emp.name} added successfully.')
        return redirect('employee_list')

    return render(request, 'attendance_app/employee_form.html', {'action': 'Add'})


@login_required
def employee_detail(request, pk):
    if not has_perm(request.user, 'employee_detail'):
        messages.error(request, 'Access denied.')
        return redirect('employee_list')

    employee = get_object_or_404(Employee, pk=pk)
    recent_attendance = Attendance.objects.filter(employee=employee).order_by('-date')[:30]
    leave_requests = LeaveRequest.objects.filter(employee=employee).order_by('-applied_on')[:10]
    activities = ActivityLog.objects.filter(
        description__icontains=employee.name
    ).order_by('-timestamp')[:30]

    # Stats
    today = timezone.localdate()
    month_start = today.replace(day=1)
    month_present = Attendance.objects.filter(employee=employee, date__gte=month_start, status='present').count()
    month_absent = Attendance.objects.filter(employee=employee, date__gte=month_start, status='absent').count()
    month_leave = Attendance.objects.filter(employee=employee, date__gte=month_start, status='leave').count()

    return render(request, 'attendance_app/employee_detail.html', {
        'employee': employee,
        'recent_attendance': recent_attendance,
        'leave_requests': leave_requests,
        'activities': activities,
        'month_present': month_present,
        'month_absent': month_absent,
        'month_leave': month_leave,
    })


@login_required
def employee_edit(request, pk):
    if not has_perm(request.user, 'employee_edit'):
        messages.error(request, 'Access denied.')
        return redirect('employee_list')

    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.name = request.POST.get('name', '').strip()
        if not employee.name:
            messages.error(request, 'Employee name is required.')
            return render(request, 'attendance_app/employee_form.html', {'employee': employee, 'action': 'Edit'})

        employee.emirates_id = request.POST.get('emirates_id') or None
        employee.mobile = request.POST.get('mobile') or None
        employee.dob = request.POST.get('dob') or None
        employee.joining_date = request.POST.get('joining_date') or None
        employee.job_title = request.POST.get('job_title') or None
        employee.country = request.POST.get('country') or None
        employee.address = request.POST.get('address') or None
        employee.description = request.POST.get('description') or None
        employee.emp_type = request.POST.get('emp_type', 'permanent')
        employee.status = request.POST.get('status', 'active')
        if request.FILES.get('photo'):
            employee.photo = request.FILES['photo']
        employee.save()

        log_activity(request.user, 'update', 'Employee', employee, f'Updated employee {employee.name}', request)
        messages.success(request, f'Employee {employee.name} updated.')
        return redirect('employee_detail', pk=pk)

    return render(request, 'attendance_app/employee_form.html', {'employee': employee, 'action': 'Edit'})


@login_required
def employee_delete(request, pk):
    if not has_perm(request.user, 'employee_delete'):
        messages.error(request, 'Access denied.')
        return redirect('employee_list')

    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        name = employee.name
        employee.delete()
        log_activity(request.user, 'delete', 'Employee', description=f'Deleted employee {name}', request=request)
        messages.success(request, f'Employee {name} deleted.')
        return redirect('employee_list')
    return render(request, 'attendance_app/employee_confirm_delete.html', {'employee': employee})


@login_required
def employee_calendar_data(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    year = int(request.GET.get('year', timezone.localdate().year))
    month = int(request.GET.get('month', timezone.localdate().month))

    attendances = Attendance.objects.filter(
        employee=employee,
        date__year=year,
        date__month=month
    )
    holidays = Holiday.objects.filter(date__year=year, date__month=month)

    events = []
    for att in attendances:
        color = {'present': '#22c55e', 'absent': '#ef4444', 'half_day': '#f59e0b',
                 'holiday': '#6366f1', 'leave': '#8b5cf6'}.get(att.status, '#6b7280')
        events.append({
            'date': att.date.isoformat(),
            'status': att.get_status_display(),
            'in_time': att.in_time.strftime('%H:%M') if att.in_time else '-',
            'out_time': att.out_time.strftime('%H:%M') if att.out_time else '-',
            'ot_hours': float(att.ot_hours),
            'total_hours': att.total_hours(),
            'color': color,
        })
    for hol in holidays:
        if not any(e['date'] == hol.date.isoformat() for e in events):
            events.append({
                'date': hol.date.isoformat(),
                'status': hol.name,
                'in_time': '-', 'out_time': '-', 'ot_hours': 0, 'total_hours': 0,
                'color': '#6366f1',
            })

    return JsonResponse({'events': events})


# ─── ATTENDANCE ───────────────────────────────────────────────────────────────

@login_required
def attendance_list(request):
    if not has_perm(request.user, 'attendance_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    today = timezone.localdate()
    date_str = request.GET.get('date', today.isoformat())
    try:
        selected_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        selected_date = today

    attendances = list(Attendance.objects.filter(date=selected_date).select_related('employee').order_by('employee__name'))
    is_holiday = Holiday.objects.filter(date=selected_date).first()

    stat_present = sum(1 for a in attendances if a.status == 'present')
    stat_absent  = sum(1 for a in attendances if a.status == 'absent')
    stat_half    = sum(1 for a in attendances if a.status == 'half_day')
    stat_leave   = sum(1 for a in attendances if a.status == 'leave')
    stat_total   = len(attendances)
    stat_total_hours = round(sum(float(a.total_hours()) for a in attendances), 1)
    stat_total_ot    = round(sum(float(a.ot_hours) for a in attendances), 1)
    attendance_rate  = round((stat_present / stat_total * 100) if stat_total else 0)

    month_qs = Attendance.objects.filter(
        date__year=selected_date.year,
        date__month=selected_date.month
    ).values("date").annotate(count=Count("id")).order_by("date")

    monthly_summary = [
        {"date": item["date"], "count": item["count"]}
        for item in month_qs
    ]

    return render(request, "attendance_app/attendance_list.html", {
        "attendances": attendances,
        "selected_date": selected_date,
        "is_holiday": is_holiday,
        "today": today,
        "yesterday": today - datetime.timedelta(days=1),
        "prev_date": selected_date - datetime.timedelta(days=1),
        "next_date": selected_date + datetime.timedelta(days=1),
        "stat_present": stat_present,
        "stat_absent": stat_absent,
        "stat_half": stat_half,
        "stat_leave": stat_leave,
        "stat_total": stat_total,
        "stat_total_hours": stat_total_hours,
        "stat_total_ot": stat_total_ot,
        "attendance_rate": attendance_rate,
        "monthly_summary": monthly_summary,
    })


@login_required
def attendance_mark(request):
    if not has_perm(request.user, 'attendance_add'):
        messages.error(request, 'Access denied.')
        return redirect('attendance_list')

    today = timezone.localdate()
    date_str = request.GET.get('date', today.isoformat())
    try:
        selected_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        selected_date = today

    active_employees = Employee.objects.filter(status='active').order_by('name')
    existing = {a.employee_id: a for a in Attendance.objects.filter(date=selected_date)}
    is_holiday = Holiday.objects.filter(date=selected_date).first()
    is_sunday = selected_date.weekday() == 6

    # Last working day data (for "copy previous day" feature)
    prev_date = selected_date - datetime.timedelta(days=1)
    prev_att = {a.employee_id: a for a in Attendance.objects.filter(date=prev_date)}

    employee_data = []
    for emp in active_employees:
        att = existing.get(emp.pk)
        prev = prev_att.get(emp.pk)
        employee_data.append({
            'employee': emp,
            'attendance': att,
            'prev_attendance': prev,
        })

    # Count already marked
    marked_count = len(existing)
    total_count = active_employees.count()

    return render(request, "attendance_app/attendance_mark.html", {
        "employee_data": employee_data,
        "selected_date": selected_date,
        "is_holiday": is_holiday,
        "is_sunday": is_sunday,
        "today": today,
        "prev_date": prev_date,
        "next_date": selected_date + datetime.timedelta(days=1),
        "marked_count": marked_count,
        "total_count": total_count,
    })


@login_required
@require_POST
def attendance_save(request):
    if not has_perm(request.user, 'attendance_add'):
        return JsonResponse({'error': 'Access denied'}, status=403)

    date_str = request.POST.get('date')
    selected_date = datetime.date.fromisoformat(date_str)
    active_employees = Employee.objects.filter(status='active')

    saved = 0
    for emp in active_employees:
        status = request.POST.get(f'status_{emp.pk}', 'absent')
        in_time_str = request.POST.get(f'in_time_{emp.pk}', '')
        out_time_str = request.POST.get(f'out_time_{emp.pk}', '')
        ot_hours = request.POST.get(f'ot_{emp.pk}', '0') or '0'

        att, created = Attendance.objects.get_or_create(
            employee=emp, date=selected_date,
            defaults={'created_by': request.user}
        )
        att.status = status
        att.in_time = datetime.time.fromisoformat(in_time_str) if in_time_str else None
        att.out_time = datetime.time.fromisoformat(out_time_str) if out_time_str else None
        try:
            att.ot_hours = float(ot_hours)
        except ValueError:
            att.ot_hours = 0
        att.save()
        saved += 1

    log_activity(request.user, 'create', 'Attendance', description=f'Marked attendance for {selected_date}', request=request)
    messages.success(request, f'Attendance saved for {selected_date} — {saved} employee(s) updated.')
    return redirect(f"/attendance/mark/?date={selected_date}")


@login_required
def attendance_edit(request, pk):
    if not has_perm(request.user, 'attendance_edit'):
        messages.error(request, 'Access denied.')
        return redirect('attendance_list')

    att = get_object_or_404(Attendance, pk=pk)
    if request.method == 'POST':
        att.status = request.POST.get('status', att.status)
        in_time_str = request.POST.get('in_time', '')
        out_time_str = request.POST.get('out_time', '')
        att.in_time = datetime.time.fromisoformat(in_time_str) if in_time_str else None
        att.out_time = datetime.time.fromisoformat(out_time_str) if out_time_str else None
        att.ot_hours = float(request.POST.get('ot_hours', 0) or 0)
        att.notes = request.POST.get('notes', '')
        att.save()
        log_activity(request.user, 'update', 'Attendance', att, f'Updated attendance for {att.employee.name} on {att.date}', request)
        messages.success(request, f'Attendance updated for {att.employee.name}.')
        return_url = request.POST.get('return_url', f'/attendance/?date={att.date}')
        return redirect(return_url)

    return render(request, 'attendance_app/attendance_edit.html', {
        'att': att,
        'return_url': request.GET.get('return_url', f'/attendance/?date={att.date}'),
    })


# ─── HOLIDAYS ─────────────────────────────────────────────────────────────────

@login_required
def holiday_list(request):
    if not has_perm(request.user, 'holiday_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    year = int(request.GET.get('year', timezone.localdate().year))
    holidays = Holiday.objects.filter(date__year=year).order_by('date')
    return render(request, 'attendance_app/holiday_list.html', {
        'holidays': holidays, 'year': year,
        'years': range(2020, timezone.localdate().year + 3),
    })


@login_required
def holiday_add(request):
    if not has_perm(request.user, 'holiday_add'):
        messages.error(request, 'Access denied.')
        return redirect('holiday_list')

    if request.method == 'POST':
        date_str = request.POST.get('date')
        name = request.POST.get('name', '').strip()
        if not date_str or not name:
            messages.error(request, 'Date and name are required.')
            return render(request, 'attendance_app/holiday_form.html')

        hol, created = Holiday.objects.get_or_create(
            date=date_str,
            defaults={'name': name, 'holiday_type': 'public'}
        )
        if not created:
            messages.warning(request, 'A holiday already exists on this date.')
        else:
            log_activity(request.user, 'create', 'Holiday', hol, f'Added holiday {name}', request)
            messages.success(request, f'Holiday {name} added.')
        return redirect('holiday_list')

    return render(request, 'attendance_app/holiday_form.html')


@login_required
def holiday_delete(request, pk):
    if not has_perm(request.user, 'holiday_add'):
        messages.error(request, 'Access denied.')
        return redirect('holiday_list')

    holiday = get_object_or_404(Holiday, pk=pk)
    if request.method == 'POST':
        name = holiday.name
        holiday.delete()
        log_activity(request.user, 'delete', 'Holiday', description=f'Deleted holiday {name}', request=request)
        messages.success(request, f'Holiday {name} deleted.')
    return redirect('holiday_list')


@login_required
def generate_sundays(request):
    if not has_perm(request.user, 'holiday_add'):
        messages.error(request, 'Access denied.')
        return redirect('holiday_list')

    year = int(request.POST.get('year', timezone.localdate().year))
    count = 0
    for month in range(1, 13):
        cal = calendar.monthcalendar(year, month)
        for week in cal:
            sunday = week[6]
            if sunday:
                date = datetime.date(year, month, sunday)
                _, created = Holiday.objects.get_or_create(
                    date=date,
                    defaults={'name': 'Sunday', 'holiday_type': 'sunday'}
                )
                if created:
                    count += 1

    log_activity(request.user, 'create', 'Holiday', description=f'Generated {count} Sundays for {year}', request=request)
    messages.success(request, f'{count} Sunday holidays generated for {year}.')
    return redirect('holiday_list')


# ─── LEAVES ──────────────────────────────────────────────────────────────────

@login_required
def leave_list(request):
    if not has_perm(request.user, 'leave_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    leaves = LeaveRequest.objects.select_related('employee', 'leave_type').all()
    status_filter = request.GET.get('status', '')
    if status_filter:
        leaves = leaves.filter(status=status_filter)
    return render(request, 'attendance_app/leave_list.html', {
        'leaves': leaves, 'status_filter': status_filter,
    })


@login_required
def leave_add(request):
    if not has_perm(request.user, 'leave_view'):
        messages.error(request, 'Access denied.')
        return redirect('leave_list')

    if request.method == 'POST':
        emp_id = request.POST.get('employee')
        lt_id = request.POST.get('leave_type')
        start = request.POST.get('start_date')
        end = request.POST.get('end_date')
        reason = request.POST.get('reason', '')

        lr = LeaveRequest.objects.create(
            employee_id=emp_id, leave_type_id=lt_id,
            start_date=start, end_date=end, reason=reason
        )
        log_activity(request.user, 'create', 'LeaveRequest', lr, f'Leave request for {lr.employee.name}', request)
        messages.success(request, 'Leave request submitted.')
        return redirect('leave_list')

    employees = Employee.objects.filter(status='active')
    leave_types = LeaveType.objects.all()
    return render(request, 'attendance_app/leave_form.html', {
        'employees': employees, 'leave_types': leave_types,
    })


@login_required
def leave_approve(request, pk):
    if not has_perm(request.user, 'leave_manage'):
        messages.error(request, 'Access denied.')
        return redirect('leave_list')

    lr = get_object_or_404(LeaveRequest, pk=pk)
    lr.status = 'approved'
    lr.reviewed_by = request.user
    lr.review_note = request.POST.get('note', '')
    lr.save()
    log_activity(request.user, 'update', 'LeaveRequest', lr, f'Approved leave for {lr.employee.name}', request)
    messages.success(request, 'Leave approved.')
    return redirect('leave_list')


@login_required
def leave_reject(request, pk):
    if not has_perm(request.user, 'leave_manage'):
        messages.error(request, 'Access denied.')
        return redirect('leave_list')

    lr = get_object_or_404(LeaveRequest, pk=pk)
    lr.status = 'rejected'
    lr.reviewed_by = request.user
    lr.review_note = request.POST.get('note', '')
    lr.save()
    log_activity(request.user, 'update', 'LeaveRequest', lr, f'Rejected leave for {lr.employee.name}', request)
    messages.success(request, 'Leave rejected.')
    return redirect('leave_list')


@login_required
def leave_type_list(request):
    leave_types = LeaveType.objects.all()
    return render(request, 'attendance_app/leave_type_list.html', {'leave_types': leave_types})


@login_required
def leave_type_add(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        days = request.POST.get('days_allowed', 0)
        desc = request.POST.get('description', '')
        if name:
            lt = LeaveType.objects.create(name=name, days_allowed=days, description=desc)
            log_activity(request.user, 'create', 'LeaveType', lt, f'Added leave type {name}', request)
            messages.success(request, f'Leave type {name} added.')
        return redirect('leave_type_list')
    return render(request, 'attendance_app/leave_type_form.html')


# ─── EXPORT ──────────────────────────────────────────────────────────────────

@login_required
def export_attendance(request):
    if not has_perm(request.user, 'export_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    today = timezone.localdate()
    return render(request, 'attendance_app/export.html', {
        'today': today,
        'months': [(i, datetime.date(2000, i, 1).strftime('%B')) for i in range(1, 13)],
        'years': range(2020, today.year + 2),
    })


@login_required
def export_get_employees(request):
    """AJAX: Return employees who attended in selected month"""
    year = int(request.GET.get('year', timezone.localdate().year))
    month = int(request.GET.get('month', timezone.localdate().month))

    employee_ids = Attendance.objects.filter(
        date__year=year, date__month=month,
        status__in=['present', 'half_day']
    ).values_list('employee_id', flat=True).distinct()

    employees = Employee.objects.filter(id__in=employee_ids).values('id', 'name')
    return JsonResponse({'employees': list(employees)})


@login_required
def export_pdf(request):
    if not has_perm(request.user, 'export_view'):
        return HttpResponse('Access denied', status=403)

    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    employee_ids = request.POST.getlist('employees')
    date_from_str = request.POST.get('date_from')
    date_to_str = request.POST.get('date_to')

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.graphics.shapes import Rect, Drawing
    except ImportError:
        return HttpResponse('ReportLab not installed. Run: pip install reportlab', status=500)

    date_from = datetime.date.fromisoformat(date_from_str) if date_from_str else datetime.date(year, month, 1)
    days_in_month = calendar.monthrange(year, month)[1]
    date_to = datetime.date.fromisoformat(date_to_str) if date_to_str else datetime.date(year, month, days_in_month)

    employees = list(Employee.objects.filter(id__in=employee_ids))
    holidays = {h.date: h.name for h in Holiday.objects.filter(date__gte=date_from, date__lte=date_to)}

    att_map = {}
    for att in Attendance.objects.filter(
        employee_id__in=employee_ids,
        date__gte=date_from, date__lte=date_to
    ).select_related('employee'):
        att_map[(att.employee_id, att.date)] = att

    cs = CompanySettings.get_settings()
    company_name    = cs.company_name
    company_address = cs.company_address
    company_phone   = cs.company_phone
    company_email   = cs.company_email
    logo_path       = cs.logo.path if cs.logo else None
    month_name      = datetime.date(year, month, 1).strftime('%B %Y')

    # ── COLORS ────────────────────────────────────────────────────────────────
    BLACK        = colors.black
    WHITE        = colors.white
    YELLOW       = colors.HexColor('#FFFF00')
    ABSENT_RED   = colors.HexColor('#CC0000')
    LEAVE_PURPLE = colors.HexColor('#7C3AED')
    HEADER_GRAY  = colors.HexColor('#E8E8E8')
    DARK_GRAY    = colors.HexColor('#404040')
    ORANGE       = colors.HexColor('#E8650A')
    LIGHT_ORANGE = colors.HexColor('#FFF3E8')

    # ── PAGE SETUP ────────────────────────────────────────────────────────────
    PAGE_W, PAGE_H = A4          # 595.27 x 841.89 pt
    L_MARGIN = R_MARGIN = 1.5 * cm
    T_MARGIN = B_MARGIN = 1.5 * cm

    # Usable width after margins
    usable_w = PAGE_W - L_MARGIN - R_MARGIN   # ~547 pt  ≈ 19.3 cm

    # 5 employees per page — compute exact col widths to fill full usable width
    EMPS_PER_PAGE = 5
    emp_chunks = [employees[i:i+EMPS_PER_PAGE] for i in range(0, len(employees), EMPS_PER_PAGE)]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=L_MARGIN, rightMargin=R_MARGIN,
        topMargin=T_MARGIN, bottomMargin=B_MARGIN
    )

    has_logo = bool(logo_path and os.path.exists(logo_path))

    story = []

    for chunk_idx, chunk in enumerate(emp_chunks):
        n = len(chunk)

        # DATE col + n*3 sub-cols — calculated to fill full page width exactly
        DATE_W = 1.45 * cm
        remaining = usable_w - DATE_W
        SUB_W = remaining / (n * 3)
        col_widths = [DATE_W] + [SUB_W] * (n * 3)

        # ── PROFESSIONAL HEADER ───────────────────────────────────────────
        header_w = sum(col_widths)

        co_name_style = ParagraphStyle('co', alignment=TA_CENTER, fontSize=16,
                                       fontName='Helvetica-Bold', leading=20,
                                       textColor=DARK_GRAY)
        co_addr_style = ParagraphStyle('addr', alignment=TA_CENTER, fontSize=8,
                                       leading=11, textColor=colors.HexColor('#666666'))
        co_sub_style  = ParagraphStyle('sub', alignment=TA_CENTER, fontSize=8,
                                       leading=11, textColor=colors.HexColor('#888888'))
        report_style  = ParagraphStyle('rep', alignment=TA_RIGHT, fontSize=9,
                                       fontName='Helvetica-Bold', leading=12, textColor=ORANGE)
        period_style  = ParagraphStyle('per', alignment=TA_RIGHT, fontSize=8,
                                       leading=11, textColor=DARK_GRAY)

        if has_logo:
            try:
                from reportlab.platypus import Image as RLImage
                logo_img = RLImage(logo_path, width=1.6*cm, height=1.6*cm)
                logo_cell = logo_img
            except Exception:
                logo_cell = ''
        else:
            # Placeholder box with "LOGO" text
            logo_cell = Paragraph(
                '<font size="7" color="#AAAAAA">LOGO</font>',
                ParagraphStyle('lp', alignment=TA_CENTER, fontSize=7,
                               textColor=colors.HexColor('#AAAAAA'))
            )

        contact_line = ''
        if company_phone:
            contact_line += f'Tel: {company_phone}'
        if company_email:
            if contact_line:
                contact_line += f'  |  Email: {company_email}'
            else:
                contact_line += f'Email: {company_email}'

        center_cell = [
            Paragraph(company_name, co_name_style),
            Paragraph(company_address, co_addr_style),
        ]
        if contact_line:
            center_cell.append(Paragraph(contact_line, co_sub_style))

        right_cell = [
            Paragraph('ATTENDANCE REPORT', report_style),
            Paragraph(month_name, period_style),
            Paragraph(
                f'{date_from.strftime("%d %b %Y")} — {date_to.strftime("%d %b %Y")}',
                ParagraphStyle('pd', alignment=TA_RIGHT, fontSize=7, leading=10,
                               textColor=colors.HexColor('#888888'))
            ),
        ]

        logo_w = 1.8 * cm
        right_w = 4.5 * cm
        center_w = header_w - logo_w - right_w

        header_tbl = Table(
            [[logo_cell, center_cell, right_cell]],
            colWidths=[logo_w, center_w, right_w]
        )
        header_tbl.setStyle(TableStyle([
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
            ('ALIGN',         (0,0),(0,0),   'CENTER'),
            ('ALIGN',         (1,0),(1,0),   'CENTER'),
            ('ALIGN',         (2,0),(2,0),   'RIGHT'),
            ('LEFTPADDING',   (0,0),(-1,-1), 4),
            ('RIGHTPADDING',  (0,0),(-1,-1), 4),
            ('TOPPADDING',    (0,0),(-1,-1), 3),
            ('BOTTOMPADDING', (0,0),(-1,-1), 3),
            # Orange left border on logo cell as accent
            ('LINEAFTER',     (0,0),(0,0),   0.5, colors.HexColor('#DDDDDD')),
            # Logo cell light bg
            ('BACKGROUND',    (0,0),(0,0),   colors.HexColor('#FFF8F3')),
            ('BOX',           (0,0),(-1,-1), 1.5, ORANGE),
            # Light background for entire header
            ('BACKGROUND',    (0,0),(-1,-1), LIGHT_ORANGE),
        ]))

        story.append(header_tbl)

        # Orange separator line
        story.append(HRFlowable(
            width=header_w, thickness=2.5, color=ORANGE,
            spaceAfter=3, spaceBefore=1
        ))

        # ── NAME ROW (row 0) ──────────────────────────────────────────────
        name_row = ['NAME']
        for emp in chunk:
            name_row.extend([emp.name, '', ''])

        # ── DATE/IN/OUT/OT SUB-HEADER (row 1) ────────────────────────────
        sub_row = ['DATE']
        for _ in chunk:
            sub_row.extend(['IN', 'OUT', 'OT'])

        # ── DATA ROWS ─────────────────────────────────────────────────────
        data_rows = []
        holiday_row_indices = []
        date_list = []

        current = date_from
        while current <= date_to:
            is_holiday = current in holidays or current.weekday() == 6
            row = [current.strftime('%d-%m-%Y')]

            for emp in chunk:
                att = att_map.get((emp.id, current))
                if att:
                    if att.status == 'absent':
                        row.extend(['A', '', ''])
                    elif att.status == 'leave':
                        row.extend(['V', '', ''])
                    elif att.status == 'holiday':
                        row.extend(['', '', ''])
                    elif att.status == 'half_day':
                        in_t  = att.in_time.strftime('%H:%M')  if att.in_time  else ''
                        out_t = att.out_time.strftime('%H:%M') if att.out_time else ''
                        ot    = str(att.ot_hours) if att.ot_hours else ''
                        row.extend([in_t, out_t, ot])
                    else:  # present
                        in_t  = att.in_time.strftime('%H:%M')  if att.in_time  else ''
                        out_t = att.out_time.strftime('%H:%M') if att.out_time else ''
                        ot    = str(att.ot_hours) if att.ot_hours else ''
                        row.extend([in_t, out_t, ot])
                else:
                    row.extend(['', '', ''] if is_holiday else ['A', '', ''])

            date_list.append(current)
            if is_holiday:
                holiday_row_indices.append(len(data_rows))
            data_rows.append(row)
            current += datetime.timedelta(days=1)

        # ── TOTALS ────────────────────────────────────────────────────────
        total_ot_row    = ['Total OT']
        holiday_ot_row  = ['Holiday OT']
        absent_cnt_row  = ['Absent']

        for emp in chunk:
            total_ot = sum(
                float(att_map[(emp.id, d)].ot_hours)
                for d in date_list
                if (emp.id, d) in att_map and att_map[(emp.id, d)].ot_hours
            )
            hol_ot = sum(
                float(att_map[(emp.id, d)].total_hours())
                for d in date_list
                if (d in holidays or d.weekday() == 6) and (emp.id, d) in att_map
            )
            # Count BOTH absent and leave days
            absent_cnt = sum(
                1 for d in date_list
                if (emp.id, d) in att_map and att_map[(emp.id, d)].status == 'absent'
            )
            leave_cnt = sum(
                1 for d in date_list
                if (emp.id, d) in att_map and att_map[(emp.id, d)].status == 'leave'
            )
            # Also count days with no record and no holiday = absent
            for d in date_list:
                if (emp.id, d) not in att_map and d not in holidays and d.weekday() != 6:
                    absent_cnt += 1

            absent_leave_fmt = f'A:{absent_cnt} / L:{leave_cnt}' if (absent_cnt or leave_cnt) else ''
            total_ot_row.extend([f'{total_ot:.1f}' if total_ot else '', '', ''])
            holiday_ot_row.extend([f'{hol_ot:.1f}' if hol_ot else '', '', ''])
            absent_cnt_row.extend([absent_leave_fmt, '', ''])

        # ── BUILD FULL TABLE ──────────────────────────────────────────────
        all_rows   = [name_row, sub_row] + data_rows + [total_ot_row, holiday_ot_row, absent_cnt_row]
        data_start = 2
        total_start = data_start + len(data_rows)

        # Row heights: header rows taller, data rows compact to fill page
        # Estimate available height for data rows
        header_h  = 2.0 * cm   # header block
        sep_h     = 0.35 * cm
        name_row_h = 0.9 * cm
        sub_row_h  = 1 * cm
        totals_h   = 0.62 * cm * 3
        available  = PAGE_H - T_MARGIN - B_MARGIN - header_h - sep_h - name_row_h - sub_row_h - totals_h
        row_h = max(0.65 * cm, min(0.55 * cm, available / max(len(data_rows), 1)))

        row_heights = [name_row_h, sub_row_h] + [row_h] * len(data_rows) + [0.62*cm, 0.62*cm, 0.62*cm]

        tbl = Table(all_rows, colWidths=col_widths, rowHeights=row_heights, repeatRows=2)

        # ── STYLES ────────────────────────────────────────────────────────
        s = [
            ('GRID',          (0,0), (-1,-1), 0.5, BLACK),
            ('FONTSIZE',      (0,0), (-1,-1), 7),
            ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING',    (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
            ('LEFTPADDING',   (0,0), (-1,-1), 1),
            ('RIGHTPADDING',  (0,0), (-1,-1), 1),

            # NAME row
            ('BACKGROUND',    (0,0), (-1,0), HEADER_GRAY),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,0), 7),

            # Sub-header row
            ('BACKGROUND',    (0,1), (-1,1), HEADER_GRAY),
            ('FONTNAME',      (0,1), (-1,1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,1), (-1,1), 7),

            # DATE column
            ('FONTNAME',      (0,2), (0,-1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,2), (0,-4), 7),

            # Totals rows
            ('BACKGROUND',    (0,total_start), (-1,-1), HEADER_GRAY),
            ('FONTNAME',      (0,total_start), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,total_start), (-1,-1), 7),
            ('ALIGN',         (0,total_start), (0,-1),  'LEFT'),
            ('LEFTPADDING',   (0,total_start), (0,-1),  3),
        ]

        # Span employee names in NAME row
        for i in range(n):
            cs = 1 + i * 3
            ce = cs + 2
            s.append(('SPAN',     (cs, 0), (ce, 0)))
            s.append(('FONTSIZE', (cs, 0), (ce, 0), 7))

        # Span totals values across 3 cols per employee
        for tr in range(3):
            ri = total_start + tr
            for i in range(n):
                cs = 1 + i * 3
                ce = cs + 2
                s.append(('SPAN', (cs, ri), (ce, ri)))

        # Yellow holiday rows — full row
        for ri in holiday_row_indices:
            s.append(('BACKGROUND', (0, ri + data_start), (-1, ri + data_start), YELLOW))

        # ABSENT = red bold, LEAVE = purple bold (per cell)
        for ri, row_data in enumerate(data_rows):
            ari = ri + data_start
            for ci, cell in enumerate(row_data):
                v = str(cell)
                if v == 'ABSENT':
                    s.append(('BACKGROUND', (ci, ari), (ci, ari), ABSENT_RED))
                    s.append(('FONTNAME',  (ci, ari), (ci, ari), 'Helvetica-Bold'))
                elif v == 'LEAVE':
                    s.append(('BACKGROUND', (ci, ari), (ci, ari), LEAVE_PURPLE))
                    s.append(('FONTNAME',  (ci, ari), (ci, ari), 'Helvetica-Bold'))

        tbl.setStyle(TableStyle(s))
        story.append(tbl)

        if chunk_idx < len(emp_chunks) - 1:
            story.append(PageBreak())

    doc.build(story)
    buffer.seek(0)

    log_activity(request.user, 'export', 'Attendance', description=f'PDF export {month_name}', request=request)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attendance_{year}_{month:02d}.pdf"'
    return response


@login_required
def export_excel(request):
    if not has_perm(request.user, 'export_view'):
        return HttpResponse('Access denied', status=403)

    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    employee_ids = request.POST.getlist('employees')
    date_from_str = request.POST.get('date_from')
    date_to_str = request.POST.get('date_to')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed. Run: pip install openpyxl', status=500)

    date_from = datetime.date.fromisoformat(date_from_str) if date_from_str else datetime.date(year, month, 1)
    days_in_month = calendar.monthrange(year, month)[1]
    date_to = datetime.date.fromisoformat(date_to_str) if date_to_str else datetime.date(year, month, days_in_month)
    employees = list(Employee.objects.filter(id__in=employee_ids))
    holidays = {h.date: h.name for h in Holiday.objects.filter(date__gte=date_from, date__lte=date_to)}

    # Pre-fetch all attendance
    att_map = {}
    for att in Attendance.objects.filter(
        employee_id__in=employee_ids,
        date__gte=date_from, date__lte=date_to
    ).select_related('employee'):
        att_map[(att.employee_id, att.date)] = att

    cs = CompanySettings.get_settings()
    company_name = cs.company_name
    company_address = cs.company_address
    month_name = datetime.date(year, month, 1).strftime('%B %Y')

    # Styles
    YELLOW_FILL  = PatternFill("solid", fgColor="FFFF00")
    HEADER_FILL  = PatternFill("solid", fgColor="D9D9D9")
    WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
    ABSENT_FILL  = PatternFill("solid", fgColor="FFFFFF")

    bold_font    = Font(bold=True, size=8)
    normal_font  = Font(size=7)
    absent_font  = Font(bold=True, size=7, color="FF0000")
    leave_font   = Font(bold=True, size=7, color="8B5CF6")
    title_font   = Font(bold=True, size=14)
    addr_font    = Font(size=9, color="444444")
    center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_center  = Alignment(horizontal='left', vertical='center')

    def thin_border():
        s = Side(style='thin', color='000000')
        return Border(left=s, right=s, top=s, bottom=s)

    BORDER = thin_border()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance {month_name}"

    # 5 employees per sheet (can add more sheets if needed)
    EMPS_PER_PAGE = 5
    emp_chunks = [employees[i:i+EMPS_PER_PAGE] for i in range(0, len(employees), EMPS_PER_PAGE)]

    current_row = 1

    for chunk_idx, chunk in enumerate(emp_chunks):
        n = len(chunk)
        # Total columns: 1 (DATE) + n*3
        total_cols = 1 + n * 3

        if chunk_idx > 0:
            current_row += 2  # spacer between chunks

        # ── COMPANY HEADER ──────────────────────────────────────────────
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        c = ws.cell(row=current_row, column=1, value=company_name)
        c.font = title_font
        c.alignment = center
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        c = ws.cell(row=current_row, column=1, value=company_address)
        c.font = addr_font
        c.alignment = center
        ws.row_dimensions[current_row].height = 14
        current_row += 1

        # ── NAME ROW ─────────────────────────────────────────────────────
        name_row_idx = current_row
        ws.cell(row=current_row, column=1, value='NAME').font = bold_font
        ws.cell(row=current_row, column=1).fill = HEADER_FILL
        ws.cell(row=current_row, column=1).alignment = center
        ws.cell(row=current_row, column=1).border = BORDER

        for i, emp in enumerate(chunk):
            c_start = 2 + i * 3
            c_end   = c_start + 2
            ws.merge_cells(start_row=current_row, start_column=c_start, end_row=current_row, end_column=c_end)
            c = ws.cell(row=current_row, column=c_start, value=emp.name)
            c.font = bold_font
            c.fill = HEADER_FILL
            c.alignment = center
            c.border = BORDER
            # borders on merged cells
            for col in range(c_start, c_end + 1):
                ws.cell(row=current_row, column=col).border = BORDER

        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # ── DATE/IN/OUT/OT SUB-HEADER ROW ────────────────────────────────
        sub_row_idx = current_row
        for col_i, val in enumerate(['DATE'] + ['IN', 'OUT', 'OT'] * n, 1):
            c = ws.cell(row=current_row, column=col_i, value=val)
            c.font = bold_font
            c.fill = HEADER_FILL
            c.alignment = center
            c.border = BORDER
        ws.row_dimensions[current_row].height = 14
        current_row += 1

        # ── DATA ROWS ─────────────────────────────────────────────────────
        date_rows_start = current_row
        date_list = []
        cur = date_from
        while cur <= date_to:
            date_list.append(cur)
            cur += datetime.timedelta(days=1)

        for cur in date_list:
            is_hol = cur in holidays or cur.weekday() == 6
            row_fill = YELLOW_FILL if is_hol else WHITE_FILL

            # DATE cell
            dc = ws.cell(row=current_row, column=1, value=cur.strftime('%d-%m-%Y'))
            dc.font = Font(bold=True, size=7)
            dc.fill = row_fill
            dc.alignment = center
            dc.border = BORDER

            for i, emp in enumerate(chunk):
                att = att_map.get((emp.id, cur))
                c_base = 2 + i * 3

                if att:
                    if att.status == 'absent':
                        vals = ['ABSENT', '', '']
                    elif att.status == 'leave':
                        vals = ['LEAVE', '', '']
                    elif att.status == 'holiday':
                        vals = ['', '', '']
                    elif att.status == 'half_day':
                        vals = [
                            att.in_time.strftime('%H:%M') if att.in_time else '',
                            att.out_time.strftime('%H:%M') if att.out_time else '',
                            str(att.ot_hours) if att.ot_hours else ''
                        ]
                    else:
                        vals = [
                            att.in_time.strftime('%H:%M') if att.in_time else '',
                            att.out_time.strftime('%H:%M') if att.out_time else '',
                            str(att.ot_hours) if att.ot_hours else ''
                        ]
                else:
                    vals = ['', '', ''] if is_hol else ['ABSENT', '', '']

                for j, v in enumerate(vals):
                    c = ws.cell(row=current_row, column=c_base + j, value=v)
                    c.fill = row_fill
                    c.alignment = center
                    c.border = BORDER
                    if v == 'ABSENT':
                        c.font = absent_font
                    elif v == 'LEAVE':
                        c.font = leave_font
                    else:
                        c.font = normal_font

            ws.row_dimensions[current_row].height = 13
            current_row += 1

        # ── TOTALS ROWS ───────────────────────────────────────────────────
        totals_labels = ['Total OT', 'Holiday OT', 'Absent / Leave']
        totals_values = [[], [], []]

        for emp in chunk:
            total_ot = sum(
                float(att_map[(emp.id, d)].ot_hours)
                for d in date_list
                if (emp.id, d) in att_map and att_map[(emp.id, d)].ot_hours
            )
            hol_ot = sum(
                float(att_map[(emp.id, d)].total_hours())
                for d in date_list
                if (d in holidays or d.weekday() == 6) and (emp.id, d) in att_map
            )
            absent_cnt = sum(
                1 for d in date_list
                if (emp.id, d) in att_map and att_map[(emp.id, d)].status == 'absent'
            )
            leave_cnt = sum(
                1 for d in date_list
                if (emp.id, d) in att_map and att_map[(emp.id, d)].status == 'leave'
            )
            # Days with no record and not a holiday = absent
            for d in date_list:
                if (emp.id, d) not in att_map and d not in holidays and d.weekday() != 6:
                    absent_cnt += 1

            absent_leave = f'A:{absent_cnt} / L:{leave_cnt}' if (absent_cnt or leave_cnt) else ''
            totals_values[0].append(f'{total_ot:.1f}' if total_ot else '')
            totals_values[1].append(f'{hol_ot:.1f}'   if hol_ot  else '')
            totals_values[2].append(absent_leave)

        for t_idx, label in enumerate(totals_labels):
            # Label cell
            lc = ws.cell(row=current_row, column=1, value=label)
            lc.font = bold_font
            lc.fill = HEADER_FILL
            lc.alignment = left_center
            lc.border = BORDER

            for i in range(n):
                c_start = 2 + i * 3
                c_end   = c_start + 2
                ws.merge_cells(start_row=current_row, start_column=c_start, end_row=current_row, end_column=c_end)
                vc = ws.cell(row=current_row, column=c_start, value=totals_values[t_idx][i])
                vc.font = bold_font
                vc.fill = HEADER_FILL
                vc.alignment = center
                for col in range(c_start, c_end + 1):
                    ws.cell(row=current_row, column=col).border = BORDER

            ws.row_dimensions[current_row].height = 14
            current_row += 1

        # ── COLUMN WIDTHS ─────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 13
        for i in range(n):
            base = 2 + i * 3
            ws.column_dimensions[get_column_letter(base)].width = 9
            ws.column_dimensions[get_column_letter(base + 1)].width = 9
            ws.column_dimensions[get_column_letter(base + 2)].width = 6

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_activity(request.user, 'export', 'Attendance', description=f'Excel export {month_name}', request=request)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_{year}_{month:02d}.xlsx"'
    return response

    wb = openpyxl.Workbook()
    ws = wb.active
    company_name = getattr(settings, 'COMPANY_NAME', 'Your Company')
    month_name = datetime.date(year, month, 1).strftime('%B %Y')
    ws.title = f"Attendance {month_name}"

    orange_fill = PatternFill("solid", fgColor="F97316")
    header_fill = PatternFill("solid", fgColor="1E293B")
    absent_fill = PatternFill("solid", fgColor="FEE2E2")
    holiday_fill = PatternFill("solid", fgColor="EDE9FE")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    bold_white = Font(bold=True, color="FFFFFF")
    bold_orange = Font(bold=True, color="F97316", size=14)
    absent_font = Font(bold=True, color="EF4444")
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = company_name
    ws['A1'].font = bold_orange
    ws.row_dimensions[1].height = 25

    ws.merge_cells('C1:E1')
    ws['C1'] = f'Attendance Report - {month_name}'
    ws['C1'].font = Font(bold=True, size=12)
    ws['C1'].alignment = center

    ws.append([])

    emp_list = list(employees)
    # Header row
    header = ['Date']
    for emp in emp_list:
        header.extend([emp.name, 'OUT', 'OT'])

    ws.append(header)
    header_row = ws.max_row
    for col_idx, val in enumerate(header, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = orange_fill if col_idx == 1 else header_fill
        cell.font = bold_white
        cell.alignment = center
        cell.border = thin

    # Merge employee name headers
    for emp_idx in range(len(emp_list)):
        start_col = 2 + emp_idx * 3
        if start_col < start_col + 1:
            pass

    ws.column_dimensions['A'].width = 14
    for emp_idx in range(len(emp_list)):
        base = 2 + emp_idx * 3
        ws.column_dimensions[get_column_letter(base)].width = 12
        ws.column_dimensions[get_column_letter(base + 1)].width = 10
        ws.column_dimensions[get_column_letter(base + 2)].width = 8

    current = date_from
    row_num = header_row + 1
    while current <= date_to:
        row_data = [current.strftime('%d %b %Y (%a)')]
        is_holiday = current in holidays
        is_sunday = current.weekday() == 6

        for emp in emp_list:
            try:
                att = Attendance.objects.get(employee=emp, date=current)
                if att.status == 'absent':
                    row_data.extend(['ABSENT', '', ''])
                elif att.status == 'half_day':
                    in_t = att.in_time.strftime('%H:%M') if att.in_time else '-'
                    out_t = att.out_time.strftime('%H:%M') if att.out_time else '-'
                    row_data.extend([f'HD {in_t}', out_t, str(att.ot_hours)])
                else:
                    in_t = att.in_time.strftime('%H:%M') if att.in_time else '-'
                    out_t = att.out_time.strftime('%H:%M') if att.out_time else '-'
                    row_data.extend([in_t, out_t, str(att.ot_hours) if att.ot_hours else ''])
            except Attendance.DoesNotExist:
                if is_holiday or is_sunday:
                    hname = holidays.get(current, 'Sunday' if is_sunday else 'Holiday')
                    row_data.extend([hname, '', ''])
                else:
                    row_data.extend(['ABSENT', '', ''])

        ws.append(row_data)
        fill = alt_fill if row_num % 2 == 0 else white_fill
        if is_holiday or is_sunday:
            fill = holiday_fill

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = fill
            cell.alignment = center
            cell.border = thin
            if str(val) == 'ABSENT':
                cell.fill = absent_fill
                cell.font = absent_font

        row_num += 1
        current += datetime.timedelta(days=1)




# ─── SETTINGS ─────────────────────────────────────────────────────────────────

@login_required
def settings_view(request):
    if not request.user.is_superuser:
        messages.error(request, 'Only administrators can access settings.')
        return redirect('dashboard')

    cs = CompanySettings.get_settings()

    if request.method == 'POST':
        tab = request.POST.get('tab', 'company')

        if tab == 'company':
            cs.company_name    = request.POST.get('company_name', '').strip() or cs.company_name
            cs.company_address = request.POST.get('company_address', '').strip()
            cs.company_phone   = request.POST.get('company_phone', '').strip()
            cs.company_email   = request.POST.get('company_email', '').strip()
            cs.company_website = request.POST.get('company_website', '').strip()
            cs.company_trn     = request.POST.get('company_trn', '').strip()
            if request.FILES.get('logo'):
                cs.logo = request.FILES['logo']
            if request.POST.get('remove_logo') == '1':
                cs.logo = None
            cs.save()
            # Also update settings.py runtime values
            from django.conf import settings as django_settings
            django_settings.COMPANY_NAME    = cs.company_name
            django_settings.COMPANY_ADDRESS = cs.company_address
            django_settings.COMPANY_PHONE   = cs.company_phone
            django_settings.COMPANY_EMAIL   = cs.company_email
            log_activity(request.user, 'update', 'Settings', description='Updated company profile', request=request)
            messages.success(request, 'Company profile saved successfully.')

        elif tab == 'work':
            in_str  = request.POST.get('default_in_time', '')
            out_str = request.POST.get('default_out_time', '')
            if in_str:
                cs.default_in_time = datetime.time.fromisoformat(in_str)
            if out_str:
                cs.default_out_time = datetime.time.fromisoformat(out_str)
            cs.work_days   = request.POST.get('work_days', cs.work_days)
            cs.weekend_day = request.POST.get('weekend_day', cs.weekend_day)
            cs.save()
            log_activity(request.user, 'update', 'Settings', description='Updated work schedule', request=request)
            messages.success(request, 'Work schedule saved successfully.')

        elif tab == 'password':
            current_pw  = request.POST.get('current_password', '')
            new_pw      = request.POST.get('new_password', '')
            confirm_pw  = request.POST.get('confirm_password', '')
            if not request.user.check_password(current_pw):
                messages.error(request, 'Current password is incorrect.')
            elif len(new_pw) < 6:
                messages.error(request, 'New password must be at least 6 characters.')
            elif new_pw != confirm_pw:
                messages.error(request, 'New passwords do not match.')
            else:
                request.user.set_password(new_pw)
                request.user.save()
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, request.user)
                log_activity(request.user, 'update', 'Auth', description='Changed admin password', request=request)
                messages.success(request, 'Password changed successfully.')

        elif tab == 'system':
            cs.timezone_name = request.POST.get('timezone_name', cs.timezone_name)
            cs.date_format   = request.POST.get('date_format', cs.date_format)
            cs.save()
            log_activity(request.user, 'update', 'Settings', description='Updated system settings', request=request)
            messages.success(request, 'System settings saved.')

        return redirect(f"/settings/?tab={tab}")

    active_tab = request.GET.get('tab', 'company')
    timezones = [
        'Asia/Dubai','Asia/Riyadh','Asia/Kuwait','Asia/Qatar','Asia/Bahrain',
        'Asia/Muscat','Asia/Karachi','Asia/Kolkata','Asia/Dhaka','Asia/Bangkok',
        'Asia/Singapore','Asia/Manila','Asia/Jakarta','Africa/Cairo','Africa/Nairobi',
        'Europe/London','Europe/Paris','Europe/Istanbul','America/New_York','America/Chicago',
        'America/Denver','America/Los_Angeles','Australia/Sydney','Pacific/Auckland',
    ]
    return render(request, 'attendance_app/settings.html', {
        'cs': cs,
        'active_tab': active_tab,
        'timezones': timezones,
    })

# ─── USER MANAGEMENT ─────────────────────────────────────────────────────────

@login_required
def user_list(request):
    if not request.user.is_superuser and not has_perm(request.user, 'user_manage'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    users = User.objects.all().select_related('custom_permissions')
    return render(request, 'attendance_app/user_list.html', {'users': users})


@login_required
def user_add(request):
    if not request.user.is_superuser and not has_perm(request.user, 'user_manage'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    all_perms = UserPermission.PERMISSION_CHOICES

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        email = request.POST.get('email', '')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        selected_perms = request.POST.getlist('permissions')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return render(request, 'attendance_app/user_form.html', {'all_perms': all_perms, 'action': 'Add'})

        user = User.objects.create_user(
            username=username, password=password,
            email=email, first_name=first_name, last_name=last_name
        )
        UserPermission.objects.create(user=user, permissions=selected_perms)
        log_activity(request.user, 'create', 'User', user, f'Created user {username}', request)
        messages.success(request, f'User {username} created.')
        return redirect('user_list')

    return render(request, 'attendance_app/user_form.html', {'all_perms': all_perms, 'action': 'Add'})


@login_required
def user_edit(request, pk):
    if not request.user.is_superuser and not has_perm(request.user, 'user_manage'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    user_obj = get_object_or_404(User, pk=pk)
    try:
        user_perm = user_obj.custom_permissions
    except UserPermission.DoesNotExist:
        user_perm = UserPermission.objects.create(user=user_obj, permissions=[])

    all_perms = UserPermission.PERMISSION_CHOICES

    if request.method == 'POST':
        user_obj.email = request.POST.get('email', '')
        user_obj.first_name = request.POST.get('first_name', '')
        user_obj.last_name = request.POST.get('last_name', '')
        new_password = request.POST.get('password', '')
        if new_password:
            user_obj.set_password(new_password)
        user_obj.save()
        user_perm.permissions = request.POST.getlist('permissions')
        user_perm.save()
        log_activity(request.user, 'update', 'User', user_obj, f'Updated user {user_obj.username}', request)
        messages.success(request, f'User {user_obj.username} updated.')
        return redirect('user_list')

    return render(request, 'attendance_app/user_form.html', {
        'user_obj': user_obj, 'user_perm': user_perm,
        'all_perms': all_perms, 'action': 'Edit',
    })


@login_required
def user_delete(request, pk):
    if not request.user.is_superuser:
        messages.error(request, 'Access denied.')
        return redirect('user_list')

    user_obj = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        if user_obj == request.user:
            messages.error(request, 'Cannot delete your own account.')
            return redirect('user_list')
        name = user_obj.username
        user_obj.delete()
        log_activity(request.user, 'delete', 'User', description=f'Deleted user {name}', request=request)
        messages.success(request, f'User {name} deleted.')
    return redirect('user_list')


# ─── ACTIVITY LOG ─────────────────────────────────────────────────────────────

@login_required
def activity_log(request):
    if not has_perm(request.user, 'activity_view'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    logs = ActivityLog.objects.select_related('user').all()
    action_filter = request.GET.get('action', '')
    model_filter = request.GET.get('model', '')
    user_filter = request.GET.get('user', '')
    date_filter = request.GET.get('date', '')

    if action_filter:
        logs = logs.filter(action=action_filter)
    if model_filter:
        logs = logs.filter(model_name__icontains=model_filter)
    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)
    if date_filter:
        logs = logs.filter(timestamp__date=date_filter)

    return render(request, 'attendance_app/activity_log.html', {
        'logs': logs[:200],
        'action_choices': ActivityLog.ACTION_CHOICES,
        'action_filter': action_filter,
        'model_filter': model_filter,
        'user_filter': user_filter,
        'date_filter': date_filter,
    })
